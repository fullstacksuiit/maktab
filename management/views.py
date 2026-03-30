
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import models, transaction
from django.db.models import Sum, Count, Q, Subquery, OuterRef, DecimalField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.template.loader import render_to_string
from datetime import date, datetime, timedelta
import calendar as cal_module
from django.conf import settings
from django.core import signing
import json
import logging
import os
import re

logger = logging.getLogger('management')

from .forms import (SignUpForm, LoginForm, CourseForm, BatchForm, StudentForm, StaffForm,
                    AttendanceFilterForm, StaffAttendanceFilterForm, FeePaymentForm, BehaviorNoteForm, SettingsForm, InviteUserForm, UserEditForm,
                    AdmissionApplicationForm, ApplicationRejectForm, EventForm,
                    LeaveTypeForm, LeaveRequestForm, LeaveRejectForm,
                    SalaryComponentForm, PayrollComponentForm, ExpenseForm)
from .models import (User, Organization, Course, Batch, Student, Staff, Attendance, StaffAttendance, FeePayment, BehaviorNote, AdmissionApplication, Event,
                     LeaveType, LeaveBalance, LeaveRequest, ensure_leave_balances,
                     PunchRecord, SalaryComponent, Payroll, PayrollComponent, create_default_salary_components, Expense)
from .decorators import role_required, admin_required, manager_or_admin_required, parent_required, internal_user_required, staff_role_required
from .indian_cities import CITY_DATA
from .hijri_dates import get_upcoming_islamic_dates
from .utils import normalize_phone


def service_worker(request):
    """Serve the service worker from root scope for PWA support."""
    sw_path = os.path.join(settings.BASE_DIR, 'management', 'static', 'management', 'sw.js')
    with open(sw_path, 'r') as f:
        return HttpResponse(f.read(), content_type='application/javascript')


def robots_txt(request):
    """Serve robots.txt for search engine crawlers."""
    host = request.build_absolute_uri('/')[:-1]
    content = f"""User-agent: *
Allow: /
Allow: /login/
Allow: /signup/
Disallow: /admin/
Disallow: /static/
Disallow: /media/

Sitemap: {host}/sitemap.xml
"""
    return HttpResponse(content.strip(), content_type='text/plain')


def sitemap_xml(request):
    """Serve sitemap.xml for search engine indexing."""
    host = request.build_absolute_uri('/')[:-1]
    content = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
    <url>
        <loc>{host}/</loc>
        <changefreq>weekly</changefreq>
        <priority>1.0</priority>
    </url>
    <url>
        <loc>{host}/login/</loc>
        <changefreq>monthly</changefreq>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>{host}/signup/</loc>
        <changefreq>monthly</changefreq>
        <priority>0.9</priority>
    </url>
    <url>
        <loc>{host}/features/</loc>
        <changefreq>monthly</changefreq>
        <priority>0.8</priority>
    </url>
</urlset>"""
    return HttpResponse(content.strip(), content_type='application/xml')


def features_page(request):
    """Public features page for SEO."""
    return render(request, 'management/features.html')


def get_org(request):
    """Helper to get the current user's organization."""
    return request.user.organization


def signup_view(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)

            # Create default leave types and salary components for the new organization
            from .models import create_default_leave_types
            create_default_leave_types(user.organization)
            create_default_salary_components(user.organization)

            # Send welcome email
            try:
                html_message = render_to_string('management/emails/welcome.html', {
                    'username': user.username,
                    'org_name': user.organization.org_name,
                })
                send_mail(
                    subject='Welcome to Maktab!',
                    message=f'Assalamu Alaikum {user.username}, your account for {user.organization.org_name} has been created successfully.',
                    from_email=None,  # uses DEFAULT_FROM_EMAIL
                    recipient_list=[user.email],
                    html_message=html_message,
                )
            except Exception as e:
                logger.error(f'Failed to send welcome email to {user.email}: {e}')

            messages.success(request, 'Account created successfully!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SignUpForm()
    return render(request, 'management/signup.html', {'form': form})


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Generate a signed remember-me token for PWA session persistence
                remember_token = signing.dumps(user.pk, salt='remember-me')
                display_name = user.first_name or username
                messages.success(request, f'Assalamu Alaikum, {display_name}!')
                if user.is_parent():
                    redirect_url = 'parent_dashboard'
                elif user.is_staff_role():
                    redirect_url = 'staff_portal'
                else:
                    redirect_url = 'dashboard'
                response = redirect(redirect_url)
                response.set_cookie(
                    'remember_token', remember_token,
                    max_age=60 * 60 * 24 * 30,  # 30 days
                    httponly=True,
                    samesite='Lax',
                )
                return response
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    return render(request, 'management/login.html', {'form': form})


def auto_login_view(request):
    """Re-authenticate user from remember-me token when session is lost."""
    if request.user.is_authenticated:
        if request.user.is_parent():
            return redirect('parent_dashboard')
        elif request.user.is_staff_role():
            return redirect('staff_portal')
        return redirect('dashboard')

    token = request.COOKIES.get('remember_token')
    if token:
        try:
            user_pk = signing.loads(token, salt='remember-me', max_age=60 * 60 * 24 * 30)
            user = User.objects.get(pk=user_pk)
            login(request, user, backend='management.backends.PhoneOrUsernameBackend')
            # Refresh the token
            new_token = signing.dumps(user.pk, salt='remember-me')
            if user.is_parent():
                redirect_url = 'parent_dashboard'
            elif user.is_staff_role():
                redirect_url = 'staff_portal'
            else:
                redirect_url = 'dashboard'
            response = redirect(redirect_url)
            response.set_cookie(
                'remember_token', new_token,
                max_age=60 * 60 * 24 * 30,
                httponly=True,
                samesite='Lax',
            )
            return response
        except (signing.BadSignature, User.DoesNotExist):
            pass

    return render(request, 'management/landing.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    response = redirect('login')
    response.delete_cookie('remember_token')
    return response


# ─── Dashboard ───────────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def dashboard_view(request):
    # Redirect staff users to their own portal
    if request.user.is_staff_role():
        return redirect('staff_portal')
    org = get_org(request)

    # Single aggregate query for all counts instead of 4 separate COUNT queries
    from django.db.models import Value, BooleanField
    counts = Organization.objects.filter(pk=org.pk).aggregate(
        total_courses=Count('courses', distinct=True),
        total_batches=Count('batches', filter=Q(batches__is_active=True), distinct=True),
        total_students=Count('students', distinct=True),
        total_staff=Count('staff_members', distinct=True),
    )
    total_courses = counts['total_courses']
    total_batches = counts['total_batches']
    total_students = counts['total_students']
    total_staff = counts['total_staff']

    # Single query for revenue (used for both total_revenue and pending calc)
    total_revenue = FeePayment.objects.filter(
        organization=org
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    # Calculate total pending fees based on batches
    total_fees_subquery = Student.objects.filter(
        organization=org
    ).annotate(
        student_total_fees=Coalesce(Sum('batches__course__fees'), 0, output_field=DecimalField())
    ).aggregate(total=Sum('student_total_fees'))['total'] or 0

    total_pending = total_fees_subquery - total_revenue

    # Single query for today's attendance with conditional count
    today = date.today()
    today_attendance = Attendance.objects.filter(organization=org, date=today).aggregate(
        today_total=Count('id'),
        today_present=Count('id', filter=Q(status__in=['Present', 'Late'])),
    )
    today_total = today_attendance['today_total']
    today_present = today_attendance['today_present']

    # Staff attendance for today
    staff_today_attendance = StaffAttendance.objects.filter(
        organization=org, date=today
    ).aggregate(
        staff_today_total=Count('id'),
        staff_today_present=Count('id', filter=Q(status__in=['Present', 'Late'])),
    )
    staff_today_total = staff_today_attendance['staff_today_total']
    staff_today_present = staff_today_attendance['staff_today_present']

    recent_students = Student.objects.filter(organization=org).prefetch_related('batches__course').order_by('-created_at')[:5]
    recent_payments = FeePayment.objects.filter(organization=org).select_related('student', 'batch__course').order_by('-created_at')[:5]

    # Get batches with student count for quick attendance
    batches_for_attendance = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers').annotate(
        student_count=Count('students')
    ).order_by('course__course_name', 'batch_name')[:6]

    # Payment method breakdown
    payment_stats = FeePayment.objects.filter(organization=org).values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    cash_total = 0
    bank_total = 0
    online_total = 0
    for stat in payment_stats:
        if stat['payment_method'] == 'Cash':
            cash_total = stat['total'] or 0
        elif stat['payment_method'] == 'Bank Transfer':
            bank_total = stat['total'] or 0
        elif stat['payment_method'] == 'Online':
            online_total = stat['total'] or 0

    # --- NEW: Monthly revenue comparison ---
    this_month_start = today.replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)

    month_revenue = FeePayment.objects.filter(
        organization=org, payment_date__gte=this_month_start
    ).aggregate(total=Sum('amount'))['total'] or 0

    last_month_revenue = FeePayment.objects.filter(
        organization=org, payment_date__gte=last_month_start, payment_date__lte=last_month_end
    ).aggregate(total=Sum('amount'))['total'] or 0

    # --- NEW: This month's enrollments ---
    month_enrollments = Student.objects.filter(
        organization=org, enrollment_date__gte=this_month_start
    ).count()

    last_month_enrollments = Student.objects.filter(
        organization=org, enrollment_date__gte=last_month_start, enrollment_date__lte=last_month_end
    ).count()

    # --- NEW: Students with highest pending fees (top 5) ---
    students_with_pending = Student.objects.filter(
        organization=org, is_orphan=False
    ).annotate(
        total_fees=Coalesce(Sum('batches__course__fees'), 0, output_field=DecimalField()),
        total_paid=Coalesce(Sum('fee_payments__amount', filter=Q(fee_payments__status='Approved')), 0, output_field=DecimalField()),
    ).annotate(
        pending_amount=models.F('total_fees') - models.F('total_paid')
    ).filter(
        pending_amount__gt=0
    ).order_by('-pending_amount')[:5]

    # --- NEW: Batch-wise attendance today ---
    batch_attendance_today = Batch.objects.filter(
        organization=org, is_active=True
    ).select_related('course').prefetch_related('teachers').annotate(
        student_count=Count('students', distinct=True),
        today_marked=Count('attendances', filter=Q(attendances__date=today), distinct=True),
        today_present=Count('attendances', filter=Q(attendances__date=today, attendances__status__in=['Present', 'Late']), distinct=True),
    ).order_by('course__course_name', 'batch_name')

    # --- NEW: Orphan students count ---
    orphan_count = Student.objects.filter(organization=org, is_orphan=True).count()

    context = {
        'total_courses': total_courses,
        'total_batches': total_batches,
        'total_students': total_students,
        'total_staff': total_staff,
        'total_revenue': total_revenue,
        'total_pending': total_pending,
        'today_total': today_total,
        'today_present': today_present,
        'recent_students': recent_students,
        'recent_payments': recent_payments,
        'batches_for_attendance': batches_for_attendance,
        'cash_total': cash_total,
        'bank_total': bank_total,
        'online_total': online_total,
        'islamic_dates': get_upcoming_islamic_dates(),
        'staff_today_total': staff_today_total,
        'staff_today_present': staff_today_present,
        # New context
        'month_revenue': month_revenue,
        'last_month_revenue': last_month_revenue,
        'month_enrollments': month_enrollments,
        'last_month_enrollments': last_month_enrollments,
        'students_with_pending': students_with_pending,
        'batch_attendance_today': batch_attendance_today,
        'orphan_count': orphan_count,
    }
    return render(request, 'management/dashboard_main.html', context)


# ─── Course Views ────────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def course_list(request):
    org = get_org(request)
    courses_qs = Course.objects.filter(organization=org).annotate(
        student_count=Count('batches__students', distinct=True)
    )

    # Server-side search
    search_query = request.GET.get('q', '').strip()
    if search_query:
        courses_qs = courses_qs.filter(
            Q(course_name__icontains=search_query) |
            Q(course_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    paginator = Paginator(courses_qs, 20)
    page_number = request.GET.get('page')
    courses = paginator.get_page(page_number)
    return render(request, 'management/course_list.html', {
        'courses': courses,
        'search_query': search_query,
    })


@login_required(login_url='login')
@internal_user_required
def course_detail(request, pk):
    org = get_org(request)
    course = get_object_or_404(Course, pk=pk, organization=org)

    batches = Batch.objects.filter(course=course, organization=org).prefetch_related(
        'teachers', 'students'
    ).annotate(
        student_count=Count('students', distinct=True),
    )

    total_students = Student.objects.filter(batches__course=course, organization=org).values('pk').distinct().count()
    active_batches = batches.filter(is_active=True).count()
    total_batches = batches.count()

    recent_payments = FeePayment.objects.filter(
        batch__course=course, organization=org
    ).select_related('student', 'batch').order_by('-payment_date')[:10]

    total_revenue = FeePayment.objects.filter(
        batch__course=course, organization=org
    ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

    context = {
        'course': course,
        'batches': batches,
        'total_students': total_students,
        'active_batches': active_batches,
        'total_batches': total_batches,
        'recent_payments': recent_payments,
        'total_revenue': total_revenue,
    }
    return render(request, 'management/course_detail.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def course_add(request):
    org = get_org(request)
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.organization = org
            course.save()
            messages.success(request, 'Course added successfully!')
            return redirect('course_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CourseForm()
    return render(request, 'management/course_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def course_edit(request, pk):
    org = get_org(request)
    course = get_object_or_404(Course, pk=pk, organization=org)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'Course updated successfully!')
            return redirect('course_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CourseForm(instance=course)
    return render(request, 'management/course_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def course_delete(request, pk):
    org = get_org(request)
    course = get_object_or_404(Course, pk=pk, organization=org)
    course_name = course.course_name
    course.delete()
    messages.success(request, f'Course "{course_name}" deleted successfully!')
    return redirect('course_list')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def course_create_ajax(request):
    """AJAX endpoint to create a course from the batch form modal."""
    org = get_org(request)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': {'__all__': ['Invalid request.']}}, status=400)

    # Supply defaults for fields not provided by the quick-add modal
    data.setdefault('duration_value', 1)
    data.setdefault('duration_unit', 'months')
    data.setdefault('description', '')

    form = CourseForm(data)
    if form.is_valid():
        course = form.save(commit=False)
        course.organization = org
        course.save()
        return JsonResponse({'success': True, 'id': course.id, 'name': str(course)})
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


# ─── Batch Views ─────────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def batch_list(request):
    org = get_org(request)
    batches_qs = Batch.objects.filter(organization=org).select_related('course').prefetch_related('teachers').annotate(
        student_count=Count('students')
    )

    # Server-side search
    search_query = request.GET.get('q', '').strip()
    if search_query:
        batches_qs = batches_qs.filter(
            Q(batch_name__icontains=search_query) |
            Q(batch_code__icontains=search_query) |
            Q(course__course_name__icontains=search_query) |
            Q(course__course_code__icontains=search_query)
        )

    # Filter by course
    course_id = request.GET.get('course')
    if course_id:
        batches_qs = batches_qs.filter(course_id=course_id)

    # Filter by active status
    status = request.GET.get('status')
    if status == 'active':
        batches_qs = batches_qs.filter(is_active=True)
    elif status == 'inactive':
        batches_qs = batches_qs.filter(is_active=False)

    # Filter by teacher
    teacher_id = request.GET.get('teacher')
    if teacher_id:
        batches_qs = batches_qs.filter(teachers__pk=teacher_id)

    courses = Course.objects.filter(organization=org)
    teachers = Staff.objects.filter(organization=org, staff_role='Teacher').order_by('first_name', 'last_name')

    # View mode: 'list' (default) or 'teacher'
    view_mode = request.GET.get('view', 'list')

    paginator = Paginator(batches_qs, 20)
    page_number = request.GET.get('page')
    batches = paginator.get_page(page_number)
    context = {
        'batches': batches,
        'courses': courses,
        'teachers': teachers,
        'search_query': search_query,
        'selected_course': course_id,
        'selected_status': status,
        'selected_teacher': teacher_id,
        'view_mode': view_mode,
    }

    if view_mode == 'teacher':
        # Build teacher-wise batch data
        teachers_with_batches = []
        for teacher in teachers:
            t_batches = batches_qs.filter(teachers=teacher).distinct()
            teachers_with_batches.append({
                'teacher': teacher,
                'batches': t_batches,
                'batch_count': t_batches.count(),
                'total_students': sum(b.student_count for b in t_batches),
            })
        # Also get unassigned batches (no teachers)
        unassigned = batches_qs.filter(teachers__isnull=True)
        context['teachers_with_batches'] = teachers_with_batches
        context['unassigned_batches'] = unassigned

    if request.headers.get('HX-Request'):
        if view_mode == 'teacher':
            return render(request, 'management/_batch_teacher_view.html', context)
        return render(request, 'management/_batch_results.html', context)
    return render(request, 'management/batch_list.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def batch_add(request):
    org = get_org(request)
    if request.method == 'POST':
        form = BatchForm(request.POST)
        form.fields['course'].queryset = Course.objects.filter(organization=org)
        form.fields['teachers'].queryset = Staff.objects.filter(organization=org, staff_role='Teacher')
        if form.is_valid():
            batch = form.save(commit=False)
            batch.organization = org
            batch.save()
            form.save_m2m()
            messages.success(request, 'Batch added successfully!')
            return redirect('batch_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BatchForm()
        form.fields['course'].queryset = Course.objects.filter(organization=org)
        form.fields['teachers'].queryset = Staff.objects.filter(organization=org, staff_role='Teacher')

        # Pre-select course if provided
        course_id = request.GET.get('course')
        if course_id:
            form.initial['course'] = course_id

    return render(request, 'management/batch_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def batch_edit(request, pk):
    org = get_org(request)
    batch = get_object_or_404(Batch, pk=pk, organization=org)
    if request.method == 'POST':
        form = BatchForm(request.POST, instance=batch)
        form.fields['course'].queryset = Course.objects.filter(organization=org)
        form.fields['teachers'].queryset = Staff.objects.filter(organization=org, staff_role='Teacher')
        if form.is_valid():
            form.save()
            messages.success(request, 'Batch updated successfully!')
            return redirect('batch_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BatchForm(instance=batch)
        form.fields['course'].queryset = Course.objects.filter(organization=org)
        form.fields['teachers'].queryset = Staff.objects.filter(organization=org, staff_role='Teacher')
    return render(request, 'management/batch_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def batch_delete(request, pk):
    org = get_org(request)
    batch = get_object_or_404(Batch, pk=pk, organization=org)
    batch_name = batch.batch_name
    batch.delete()
    messages.success(request, f'Batch "{batch_name}" deleted successfully!')
    return redirect('batch_list')


@login_required(login_url='login')
@internal_user_required
def batch_detail(request, pk):
    org = get_org(request)
    batch = get_object_or_404(
        Batch.objects.select_related('course').prefetch_related('teachers', 'students'),
        pk=pk, organization=org
    )

    students = batch.students.all()
    teachers = batch.teachers.all()

    attendance_stats = batch.attendances.aggregate(
        total=Count('id'),
        present=Count('id', filter=Q(status='Present')),
        absent=Count('id', filter=Q(status='Absent')),
        late=Count('id', filter=Q(status='Late')),
        excused=Count('id', filter=Q(status='Excused')),
    )
    total = attendance_stats['total'] or 0
    present = (attendance_stats['present'] or 0) + (attendance_stats['late'] or 0)
    attendance_pct = round((present / total) * 100) if total > 0 else 0

    context = {
        'batch': batch,
        'students': students,
        'teachers': teachers,
        'attendance_stats': attendance_stats,
        'attendance_pct': attendance_pct,
    }
    return render(request, 'management/batch_detail.html', context)


# ─── Timetable Views ─────────────────────────────────────────────────────────

DAYS_MAP = {
    'weekdays': [0, 1, 2, 3, 4],
    'weekend': [5, 6],
    'mwf': [0, 2, 4],
    'tts': [1, 3, 5],
    'daily': [0, 1, 2, 3, 4, 5, 6],
    'custom': [0, 1, 2, 3, 4, 5, 6],
}

REVERSE_DAYS_MAP = {
    (0, 1, 2, 3, 4): 'weekdays',
    (5, 6): 'weekend',
    (0, 2, 4): 'mwf',
    (1, 3, 5): 'tts',
    (0, 1, 2, 3, 4, 5, 6): 'daily',
}


@login_required(login_url='login')
@internal_user_required
def batch_timetable(request):
    org = get_org(request)
    batches = Batch.objects.filter(
        organization=org, is_active=True, start_time__isnull=False, end_time__isnull=False
    ).select_related('course').prefetch_related('teachers').annotate(student_count=Count('students'))

    course_colors = {}
    palette = [
        '#0d6b4e', '#2563eb', '#9333ea', '#dc2626', '#ea580c',
        '#0891b2', '#4f46e5', '#c026d3', '#059669', '#d97706',
    ]
    color_idx = 0
    for b in batches:
        cid = b.course_id
        if cid not in course_colors:
            course_colors[cid] = palette[color_idx % len(palette)]
            color_idx += 1

    batch_data = []
    for b in batches:
        days = DAYS_MAP.get(b.days, [])
        teachers = ', '.join([f"{t.first_name} {t.last_name}".strip() for t in b.teachers.all()])
        batch_data.append({
            'id': b.pk,
            'name': b.batch_name,
            'code': b.batch_code,
            'course': b.course.course_name,
            'course_code': b.course.course_code,
            'teachers': teachers,
            'student_count': b.student_count,
            'days': days,
            'start_time': b.start_time.strftime('%H:%M'),
            'end_time': b.end_time.strftime('%H:%M'),
            'color': course_colors[b.course_id],
        })

    courses = Course.objects.filter(organization=org)
    context = {
        'batch_data_json': json.dumps(batch_data),
        'courses': courses,
        'course_colors_json': json.dumps({
            c.course_name: course_colors.get(c.pk, '#6b7280') for c in courses
        }),
    }
    return render(request, 'management/batch_timetable.html', context)


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def batch_schedule_update(request):
    org = get_org(request)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    batch_id = data.get('batch_id')
    new_day = data.get('day')  # 0-6 (Mon-Sun)
    new_start = data.get('start_time')  # "HH:MM"
    new_end = data.get('end_time')  # "HH:MM"

    if batch_id is None or new_day is None or new_start is None or new_end is None:
        return JsonResponse({'error': 'Missing fields'}, status=400)

    batch = get_object_or_404(Batch, pk=batch_id, organization=org)

    # Parse times
    try:
        start_time = datetime.strptime(new_start, '%H:%M').time()
        end_time = datetime.strptime(new_end, '%H:%M').time()
    except ValueError:
        return JsonResponse({'error': 'Invalid time format'}, status=400)

    if start_time >= end_time:
        return JsonResponse({'error': 'Start time must be before end time'}, status=400)

    # Determine the new days value
    old_days_list = DAYS_MAP.get(batch.days, [])
    if new_day in old_days_list:
        # Day hasn't changed, just update times
        new_days_list = old_days_list
    else:
        # Moved to a different day — update days to include only the new day
        new_days_list = [new_day]

    # Find matching days choice
    new_days_tuple = tuple(sorted(new_days_list))
    new_days_value = REVERSE_DAYS_MAP.get(new_days_tuple, 'custom')

    batch.start_time = start_time
    batch.end_time = end_time
    batch.days = new_days_value
    batch.save()

    return JsonResponse({
        'success': True,
        'batch_id': batch.pk,
        'days': new_days_value,
        'start_time': start_time.strftime('%H:%M'),
        'end_time': end_time.strftime('%H:%M'),
    })


# ─── Student Views ───────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def student_list(request):
    org = get_org(request)
    active_tab = request.GET.get('tab', 'students')

    # --- Students tab data ---
    students_qs = Student.objects.filter(organization=org).prefetch_related('batches__course').annotate(
        total_fees=Sum('batches__course__fees'),
    )
    student_search = request.GET.get('q', '').strip() if active_tab == 'students' else ''
    if student_search:
        students_qs = students_qs.filter(
            Q(full_name__icontains=student_search) |
            Q(student_id__icontains=student_search) |
            Q(email__icontains=student_search) |
            Q(phone__icontains=student_search)
        )
    paginator = Paginator(students_qs, 20)
    students = paginator.get_page(request.GET.get('page') if active_tab == 'students' else 1)

    # --- Applications tab data ---
    applications_qs = AdmissionApplication.objects.filter(organization=org)
    app_status_filter = request.GET.get('status', '').strip() if active_tab == 'applications' else ''
    if app_status_filter in ('pending', 'accepted', 'rejected'):
        applications_qs = applications_qs.filter(status=app_status_filter)
    app_search = request.GET.get('q', '').strip() if active_tab == 'applications' else ''
    if app_search:
        applications_qs = applications_qs.filter(
            Q(first_name__icontains=app_search) |
            Q(last_name__icontains=app_search) |
            Q(phone__icontains=app_search) |
            Q(email__icontains=app_search)
        )
    app_status_counts = AdmissionApplication.objects.filter(organization=org).values('status').annotate(count=Count('id'))
    app_counts = {item['status']: item['count'] for item in app_status_counts}
    app_paginator = Paginator(applications_qs, 20)
    applications = app_paginator.get_page(request.GET.get('page') if active_tab == 'applications' else 1)

    # Batches for bulk assign
    batches = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    context = {
        'active_tab': active_tab,
        'students': students,
        'search_query': student_search if active_tab == 'students' else app_search,
        'batches': batches,
        # Application context
        'applications': applications,
        'app_status_filter': app_status_filter,
        'app_pending_count': app_counts.get('pending', 0),
        'app_accepted_count': app_counts.get('accepted', 0),
        'app_rejected_count': app_counts.get('rejected', 0),
        'app_total_count': sum(app_counts.values()),
    }
    if request.headers.get('HX-Request'):
        if active_tab == 'applications':
            return render(request, 'management/_application_results.html', context)
        return render(request, 'management/_student_results.html', context)
    return render(request, 'management/student_list.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def student_add(request):
    org = get_org(request)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        form.fields['batches'].queryset = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')
        if form.is_valid():
            student = form.save(commit=False)
            student.organization = org
            student.save()
            form.save_m2m()
            messages.success(request, 'Student added successfully!')
            return redirect('student_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm()
        form.fields['batches'].queryset = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')
    return render(request, 'management/student_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def student_edit(request, uuid):
    org = get_org(request)
    student = get_object_or_404(Student, uuid=uuid, organization=org)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        form.fields['batches'].queryset = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully!')
            return redirect('student_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentForm(instance=student)
        form.fields['batches'].queryset = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')
    return render(request, 'management/student_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def student_delete(request, uuid):
    org = get_org(request)
    student = get_object_or_404(Student, uuid=uuid, organization=org)
    student_name = student.full_name
    student.delete()
    messages.success(request, f'Student "{student_name}" deleted successfully!')
    return redirect('student_list')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def student_bulk_delete(request):
    org = get_org(request)
    uuids = request.POST.getlist('selected_students')
    if not uuids:
        messages.warning(request, 'No students selected.')
        return redirect('student_list')
    students = Student.objects.filter(uuid__in=uuids, organization=org)
    count = students.count()
    students.delete()
    messages.success(request, f'{count} student(s) deleted successfully!')
    return redirect('student_list')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def student_bulk_assign_batch(request):
    org = get_org(request)
    uuids = request.POST.getlist('selected_students')
    batch_id = request.POST.get('batch')
    if not uuids or not batch_id:
        messages.warning(request, 'No students or batch selected.')
        return redirect('student_list')
    batch = get_object_or_404(Batch, pk=batch_id, organization=org)
    students = Student.objects.filter(uuid__in=uuids, organization=org)
    count = 0
    for student in students:
        if not student.batches.filter(pk=batch.pk).exists():
            student.batches.add(batch)
            count += 1
    messages.success(request, f'{count} student(s) assigned to {batch.course.course_code} - {batch.batch_name}.')
    return redirect('student_list')


@login_required(login_url='login')
@manager_or_admin_required
def student_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    org = get_org(request)
    students = Student.objects.filter(organization=org).prefetch_related('batches__course', 'fee_payments').order_by('student_id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = [
        'Student ID', 'Full Name', 'Guardian Name', 'Contact 1', 'Contact 2',
        'Address', 'City', 'Enrollment Date', 'Date of Birth', 'Gender',
        'Email', 'Batch', 'Discount', 'Last Paid',
    ]

    # Style header row
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D6B4E', end_color='0D6B4E', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, student in enumerate(students, 2):
        batches_str = ', '.join(
            b.batch_code for b in student.batches.all() if b.batch_code
        )
        discount = float(student.discount_value) if student.discount_value else ''
        # Get last paid month
        last_payment = student.fee_payments.filter(status='Approved').order_by('-fee_month_to').first()
        last_paid = last_payment.fee_month_to.strftime('%B %Y').lower() if last_payment and last_payment.fee_month_to else ''
        row_data = [
            student.student_id,
            student.full_name,
            student.guardian_name,
            student.phone,
            student.guardian_phone,
            student.address,
            student.city,
            student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
            student.gender,
            student.email,
            batches_str,
            discount,
            last_paid,
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="students_{org.org_name}_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
@manager_or_admin_required
def student_import_excel(request):
    import openpyxl

    org = get_org(request)
    batches = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file.')
            return redirect('student_import_excel')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
            return redirect('student_import_excel')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            if not rows:
                messages.warning(request, 'The Excel file has no data rows.')
                return redirect('student_import_excel')

            gender_map = {'male': 'M', 'female': 'F', 'other': 'O', 'm': 'M', 'f': 'F', 'o': 'O'}

            # Build batch lookup with multiple keys for flexible matching
            batch_lookup = {}
            for b in batches:
                # Primary key: batch_code
                if b.batch_code:
                    batch_lookup[b.batch_code.strip().lower()] = b
                # Also match by batch_name alone
                batch_lookup[b.batch_name.strip().lower()] = b
                # Also match by "COURSE_CODE - BATCH_NAME"
                batch_lookup[f'{b.course.course_code} - {b.batch_name}'.strip().lower()] = b
                # Also match by course_code alone
                batch_lookup[b.course.course_code.strip().lower()] = b

            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []

            for row_idx, row in enumerate(rows, start=2):
                # Skip completely empty rows
                if not any(row):
                    continue

                try:
                    student_id = str(row[0] or '').strip()
                    full_name = str(row[1] or '').strip()
                    guardian_name = str(row[2] or '').strip()
                    phone = str(row[3] or '').strip()
                    guardian_phone = str(row[4] or '').strip()
                    address = str(row[5] or '').strip()
                    city = str(row[6] or '').strip() or 'Rourkela'
                    enrollment_raw = row[7]
                    dob_raw = row[8]
                    gender_raw = str(row[9] or '').strip().lower()
                    email = str(row[10] or '').strip()
                    batches_raw = str(row[11] or '').strip()
                    discount_raw = row[12] if len(row) > 12 else None
                    last_paid_raw = row[13] if len(row) > 13 else None

                    # Parse discount (fixed amount)
                    discount_value = 0
                    if discount_raw is not None and str(discount_raw).strip():
                        try:
                            discount_value = float(discount_raw)
                        except (ValueError, TypeError):
                            discount_value = 0

                    # Parse last paid month
                    last_paid_month = None
                    if last_paid_raw is not None and str(last_paid_raw).strip():
                        lp_str = str(last_paid_raw).strip()
                        for fmt in ('%B %Y', '%b %Y', '%B', '%b'):
                            try:
                                parsed = datetime.strptime(lp_str, fmt)
                                if '%Y' not in fmt:
                                    parsed = parsed.replace(year=date.today().year)
                                last_paid_month = parsed.date().replace(day=1)
                                break
                            except ValueError:
                                continue

                    if not full_name:
                        errors.append(f'Row {row_idx}: Full Name is required.')
                        skipped_count += 1
                        continue

                    if not phone:
                        errors.append(f'Row {row_idx}: Contact 1 is required.')
                        skipped_count += 1
                        continue

                    # Parse date of birth
                    dob = None
                    if dob_raw:
                        if isinstance(dob_raw, datetime):
                            dob = dob_raw.date()
                        elif isinstance(dob_raw, date):
                            dob = dob_raw
                        else:
                            try:
                                dob = datetime.strptime(str(dob_raw).strip(), '%Y-%m-%d').date()
                            except ValueError:
                                pass

                    # Parse gender
                    gender = gender_map.get(gender_raw, 'M')

                    # Parse enrollment date
                    enrollment_date = None
                    if enrollment_raw:
                        if isinstance(enrollment_raw, datetime):
                            enrollment_date = enrollment_raw.date()
                        elif isinstance(enrollment_raw, date):
                            enrollment_date = enrollment_raw
                        else:
                            try:
                                enrollment_date = datetime.strptime(str(enrollment_raw).strip(), '%Y-%m-%d').date()
                            except ValueError:
                                pass
                    if not enrollment_date:
                        enrollment_date = date.today()

                    # Update existing student or create new one
                    existing = None
                    if student_id:
                        existing = Student.objects.filter(organization=org, student_id=student_id).first()

                    if existing:
                        existing.full_name = full_name
                        existing.guardian_name = guardian_name
                        existing.phone = phone
                        existing.guardian_phone = guardian_phone
                        existing.address = address
                        existing.city = city
                        existing.email = email
                        existing.date_of_birth = dob
                        existing.gender = gender
                        existing.enrollment_date = enrollment_date
                        if discount_value > 0:
                            existing.discount_type = 'fixed'
                            existing.discount_value = discount_value
                        existing.save()
                        student = existing

                        # Replace batches with what's in the file (only if column has data)
                        if batches_raw:
                            new_batches = []
                            for batch_entry in batches_raw.split(','):
                                batch_key = batch_entry.strip().lower()
                                if batch_key and batch_key in batch_lookup:
                                    new_batches.append(batch_lookup[batch_key])
                            if new_batches:
                                student.batches.set(new_batches)
                            else:
                                errors.append(f'Row {row_idx}: Batch "{batches_raw}" not found, batches unchanged.')

                        # Create fee payments from enrollment to last_paid_month
                        if last_paid_month:
                            enrollment_start = (student.enrollment_date or date.today()).replace(day=1)
                            if last_paid_month >= enrollment_start:
                                batches_qs = student.batches.select_related('course').all()
                                num_batches = batches_qs.count()
                                for batch in batches_qs:
                                    if not FeePayment.objects.filter(student=student, batch=batch, status='Approved', organization=org).exists():
                                        fee = float(batch.course.fees or 0)
                                        if student.is_orphan:
                                            fee = 0
                                        elif student.discount_value and student.discount_value > 0 and num_batches > 0:
                                            if student.discount_type == 'percentage':
                                                fee -= fee * float(student.discount_value) / 100
                                            else:
                                                fee -= float(student.discount_value) / num_batches
                                            fee = max(fee, 0)
                                        months_count = (last_paid_month.year - enrollment_start.year) * 12 + (last_paid_month.month - enrollment_start.month) + 1
                                        FeePayment.objects.create(
                                            student=student, batch=batch,
                                            amount=fee * months_count,
                                            fee_month_from=enrollment_start,
                                            fee_month_to=last_paid_month,
                                            payment_date=date.today(),
                                            payment_method='Cash',
                                            status='Approved',
                                            notes='Imported from Excel',
                                            organization=org,
                                        )

                        updated_count += 1
                    else:
                        student = Student(
                            student_id=student_id,
                            full_name=full_name,
                            guardian_name=guardian_name,
                            phone=phone,
                            guardian_phone=guardian_phone,
                            address=address,
                            city=city,
                            email=email,
                            date_of_birth=dob,
                            gender=gender,
                            enrollment_date=enrollment_date,
                            discount_type='fixed' if discount_value > 0 else '',
                            discount_value=discount_value,
                            organization=org,
                        )
                        student.save()

                        # Assign batches for new students
                        if batches_raw:
                            for batch_entry in batches_raw.split(','):
                                batch_key = batch_entry.strip().lower()
                                if batch_key in batch_lookup:
                                    student.batches.add(batch_lookup[batch_key])

                        # Create fee payments from enrollment to last_paid_month
                        if last_paid_month:
                            enrollment_start = (student.enrollment_date or date.today()).replace(day=1)
                            if last_paid_month >= enrollment_start:
                                batches_qs = student.batches.select_related('course').all()
                                num_batches = batches_qs.count()
                                for batch in batches_qs:
                                    fee = float(batch.course.fees or 0)
                                    if student.is_orphan:
                                        fee = 0
                                    elif student.discount_value and student.discount_value > 0 and num_batches > 0:
                                        if student.discount_type == 'percentage':
                                            fee -= fee * float(student.discount_value) / 100
                                        else:
                                            fee -= float(student.discount_value) / num_batches
                                        fee = max(fee, 0)
                                    months_count = (last_paid_month.year - enrollment_start.year) * 12 + (last_paid_month.month - enrollment_start.month) + 1
                                    FeePayment.objects.create(
                                        student=student, batch=batch,
                                        amount=fee * months_count,
                                        fee_month_from=enrollment_start,
                                        fee_month_to=last_paid_month,
                                        payment_date=date.today(),
                                        payment_method='Cash',
                                        status='Approved',
                                        notes='Imported from Excel',
                                        organization=org,
                                    )

                        created_count += 1

                except Exception as e:
                    errors.append(f'Row {row_idx}: {str(e)}')
                    skipped_count += 1

            parts = []
            if created_count:
                parts.append(f'{created_count} created')
            if updated_count:
                parts.append(f'{updated_count} updated')
            if parts:
                messages.success(request, f'Students imported: {", ".join(parts)}.')
            if skipped_count:
                messages.warning(request, f'{skipped_count} row(s) skipped. ' + ' | '.join(errors[:5]))
            if not created_count and not updated_count and not skipped_count:
                messages.info(request, 'No students found in the file.')

        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')

        return redirect('student_list')

    return render(request, 'management/student_import.html', {'batches': batches})


@login_required(login_url='login')
@internal_user_required
def student_detail(request, uuid):
    org = get_org(request)
    student = get_object_or_404(
        Student.objects.prefetch_related('batches__course'),
        uuid=uuid, organization=org
    )
    attendances = Attendance.objects.filter(
        student=student, organization=org
    ).select_related('batch__course').order_by('-date')[:20]
    fee_payments = FeePayment.objects.filter(
        student=student, organization=org
    ).select_related('batch__course').order_by('-payment_date')
    behavior_notes = BehaviorNote.objects.filter(
        student=student, organization=org
    ).select_related('noted_by')

    total_fees = student.get_effective_fee()
    total_paid = student.get_total_paid()

    # Build month-wise dues status per batch
    today = date.today()
    dues_by_batch = []
    # Pre-fetch all approved payments once and group by batch_id to avoid N+1
    approved_payments_list = list(fee_payments.filter(status='Approved'))
    payments_by_batch = {}
    for payment in approved_payments_list:
        payments_by_batch.setdefault(payment.batch_id, []).append(payment)

    for batch in student.batches.select_related('course').all():
        # Collect all paid months for this batch from pre-grouped data
        paid_months = set()
        for payment in payments_by_batch.get(batch.pk, []):
            for m in payment.fee_months_list:
                paid_months.add((m.year, m.month))

        # Generate expected months: from enrollment (or batch start) to current month
        start = student.enrollment_date.replace(day=1) if student.enrollment_date else today.replace(day=1)
        end = today.replace(day=1)
        current = start
        months = []
        while current <= end:
            status = 'paid' if (current.year, current.month) in paid_months else 'pending'
            months.append({
                'date': current,
                'label': current.strftime('%b %Y'),
                'short': current.strftime('%b'),
                'year': current.year,
                'status': status,
            })
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        pending_count = sum(1 for m in months if m['status'] == 'pending')
        paid_count = sum(1 for m in months if m['status'] == 'paid')

        # Apply student discount to monthly fee
        base_fee = batch.course.fees or 0
        if student.is_orphan:
            monthly_fee = 0
        elif student.discount_value and student.discount_value > 0:
            if student.discount_type == 'percentage':
                monthly_fee = base_fee - (base_fee * student.discount_value / 100)
            else:
                monthly_fee = max(base_fee - student.discount_value, 0)
        else:
            monthly_fee = base_fee

        dues_by_batch.append({
            'batch': batch,
            'months': months,
            'pending_count': pending_count,
            'paid_count': paid_count,
            'monthly_fee': monthly_fee,
            'original_fee': base_fee,
            'pending_amount': pending_count * monthly_fee,
        })

    # Compute attendance percentage in a single query instead of model method
    att_result = Attendance.objects.filter(
        student=student, organization=org
    ).aggregate(
        total=Count('id'),
        present=Count('id', filter=Q(status__in=['Present', 'Late']))
    )
    att_total = att_result['total'] or 0
    attendance_percentage = round((att_result['present'] / att_total) * 100, 1) if att_total > 0 else 0

    context = {
        'student': student,
        'attendances': attendances,
        'fee_payments': fee_payments,
        'behavior_notes': behavior_notes,
        'attendance_percentage': attendance_percentage,
        'total_paid': total_paid,
        'pending_fees': total_fees + student.opening_balance - total_paid,
        'dues_by_batch': dues_by_batch,
    }
    return render(request, 'management/student_detail.html', context)


@login_required(login_url='login')
@internal_user_required
def student_fee_history(request, uuid):
    org = get_org(request)
    student = get_object_or_404(Student.objects.prefetch_related('batches__course'), uuid=uuid, organization=org)
    payments = FeePayment.objects.filter(student=student, organization=org).select_related('batch__course')

    total_fees = student.get_effective_fee()
    total_paid = student.get_total_paid()

    context = {
        'student': student,
        'payments': payments,
        'total_paid': total_paid,
        'total_fees': total_fees,
        'pending_fees': student.get_pending_fees(),
    }
    return render(request, 'management/student_fee_history.html', context)


# ─── Staff Views ─────────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def staff_list(request):
    org = get_org(request)
    staff_qs = Staff.objects.filter(organization=org)

    # Server-side search
    search_query = request.GET.get('q', '').strip()
    if search_query:
        staff_qs = staff_qs.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(staff_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(staff_role__icontains=search_query) |
            Q(department__icontains=search_query)
        )

    paginator = Paginator(staff_qs, 20)
    page_number = request.GET.get('page')
    staff_members = paginator.get_page(page_number)
    return render(request, 'management/staff_list.html', {
        'staff_members': staff_members,
        'search_query': search_query,
    })


@login_required(login_url='login')
@manager_or_admin_required
def staff_add(request):
    org = get_org(request)
    leave_types = LeaveType.objects.filter(organization=org)
    if request.method == 'POST':
        form = StaffForm(request.POST, request.FILES, organization=org)
        if form.is_valid():
            staff = form.save(commit=False)
            staff.organization = org
            if request.POST.get('remove_photo') == '1':
                staff.photo = None
            staff.save()
            # Auto-create a User account for staff login (staff_id + phone)
            if not hasattr(staff, 'user_account') or staff.user_account is None:
                from .models import User as UserModel
                if not UserModel.objects.filter(username=staff.staff_id).exists():
                    user_account = UserModel(
                        username=staff.staff_id,
                        first_name=staff.first_name,
                        last_name=staff.last_name,
                        email=staff.email,
                        role='staff',
                        organization=org,
                        staff_profile=staff,
                    )
                    user_account.set_password(normalize_phone(staff.phone))
                    user_account.save()
            # Create leave balances with custom allocations
            current_year = date.today().year
            for lt in leave_types:
                alloc_value = request.POST.get(f'leave_alloc_{lt.id}', '').strip()
                if alloc_value != '':
                    try:
                        days = int(alloc_value)
                    except (ValueError, TypeError):
                        days = lt.days_per_year
                    yearly = days * 12 if lt.period == 'monthly' else days
                else:
                    yearly = lt.yearly_allocation
                LeaveBalance.objects.create(
                    organization=org, staff=staff, leave_type=lt,
                    year=current_year, allocated=yearly
                )
            messages.success(request, 'Staff member added successfully!')
            return redirect('staff_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StaffForm(organization=org)

    # Build leave allocation data for template
    leave_alloc_data = []
    for lt in leave_types:
        posted_val = request.POST.get(f'leave_alloc_{lt.id}', '') if request.method == 'POST' else ''
        leave_alloc_data.append({
            'id': lt.id,
            'name': lt.name,
            'code': lt.code,
            'period': lt.get_period_display(),
            'default_days': lt.days_per_year,
            'value': posted_val if posted_val != '' else lt.days_per_year,
        })

    return render(request, 'management/staff_form.html', {
        'form': form, 'action': 'Add', 'leave_alloc_data': leave_alloc_data,
    })


@login_required(login_url='login')
@manager_or_admin_required
def staff_edit(request, pk):
    org = get_org(request)
    staff = get_object_or_404(Staff, pk=pk, organization=org)
    leave_types = LeaveType.objects.filter(organization=org)
    current_year = date.today().year
    if request.method == 'POST':
        form = StaffForm(request.POST, request.FILES, instance=staff, organization=org)
        if form.is_valid():
            staff = form.save(commit=False)
            if request.POST.get('remove_photo') == '1':
                if staff.photo:
                    staff.photo.delete(save=False)
                staff.photo = None
            staff.save()
            # Sync linked User account details (but NOT password - only sync name/email/username)
            try:
                user_account = staff.user_account
                if user_account:
                    user_account.username = staff.staff_id
                    user_account.first_name = staff.first_name
                    user_account.last_name = staff.last_name
                    user_account.email = staff.email
                    user_account.save()
            except User.DoesNotExist:
                pass
            # Update leave balances with custom allocations
            for lt in leave_types:
                alloc_value = request.POST.get(f'leave_alloc_{lt.id}', '').strip()
                if alloc_value != '':
                    try:
                        days = int(alloc_value)
                    except (ValueError, TypeError):
                        days = lt.days_per_year
                    yearly = days * 12 if lt.period == 'monthly' else days
                else:
                    yearly = lt.yearly_allocation
                LeaveBalance.objects.update_or_create(
                    organization=org, staff=staff, leave_type=lt, year=current_year,
                    defaults={'allocated': yearly}
                )
            messages.success(request, 'Staff member updated successfully!')
            return redirect('staff_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StaffForm(instance=staff, organization=org)

    # Build leave allocation data for template
    existing_balances = {
        lb.leave_type_id: lb for lb in
        LeaveBalance.objects.filter(organization=org, staff=staff, year=current_year)
    }
    leave_alloc_data = []
    for lt in leave_types:
        existing = existing_balances.get(lt.id)
        if request.method == 'POST':
            val = request.POST.get(f'leave_alloc_{lt.id}', '')
            display_val = val if val != '' else lt.days_per_year
        elif existing:
            # Show per-period value (reverse the yearly conversion)
            display_val = int(float(existing.allocated) / 12) if lt.period == 'monthly' else int(float(existing.allocated))
        else:
            display_val = lt.days_per_year
        leave_alloc_data.append({
            'id': lt.id,
            'name': lt.name,
            'code': lt.code,
            'period': lt.get_period_display(),
            'default_days': lt.days_per_year,
            'value': display_val,
        })

    return render(request, 'management/staff_form.html', {
        'form': form, 'action': 'Edit', 'leave_alloc_data': leave_alloc_data,
    })


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def staff_delete(request, pk):
    org = get_org(request)
    staff = get_object_or_404(Staff, pk=pk, organization=org)
    staff_name = f"{staff.first_name} {staff.last_name}"
    staff.delete()
    messages.success(request, f'Staff member "{staff_name}" deleted successfully!')
    return redirect('staff_list')


@login_required(login_url='login')
@internal_user_required
def staff_detail(request, pk):
    org = get_org(request)
    staff = get_object_or_404(Staff, pk=pk, organization=org)
    from datetime import date
    current_year = date.today().year
    ensure_leave_balances(staff, current_year)
    leave_balances = LeaveBalance.objects.filter(
        organization=org, staff=staff, year=current_year
    ).select_related('leave_type')
    return render(request, 'management/staff_detail.html', {
        'staff': staff,
        'leave_balances': leave_balances,
    })


# ─── Attendance Views ────────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def attendance_list(request):
    org = get_org(request)
    attendances = Attendance.objects.filter(organization=org).select_related('student', 'batch__course', 'marked_by')

    batch_id = request.GET.get('batch')
    filter_date = request.GET.get('date')
    if batch_id:
        attendances = attendances.filter(batch_id=batch_id)
    if filter_date:
        attendances = attendances.filter(date=filter_date)

    batches = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    paginator = Paginator(attendances, 50)
    page_number = request.GET.get('page')
    attendances_page = paginator.get_page(page_number)

    context = {
        'attendances': attendances_page,
        'batches': batches,
        'selected_batch': batch_id,
        'selected_date': filter_date,
    }
    return render(request, 'management/attendance_list.html', context)


@login_required(login_url='login')
@internal_user_required
def attendance_mark(request):
    org = get_org(request)
    if request.method == 'POST':
        batch_id = request.POST.get('batch')
        attendance_date = request.POST.get('date')
        batch = get_object_or_404(Batch, pk=batch_id, organization=org)

        students = list(batch.students.filter(organization=org))
        existing = {
            a.student_id: a for a in Attendance.objects.filter(
                date=attendance_date, batch=batch, organization=org,
                student__in=students
            )
        }
        to_create = []
        to_update = []
        for student in students:
            status = request.POST.get(f'status_{student.pk}', 'Absent')
            notes = request.POST.get(f'notes_{student.pk}', '')
            if student.pk in existing:
                att = existing[student.pk]
                att.status = status
                att.marked_by = request.user
                att.notes = notes
                to_update.append(att)
            else:
                to_create.append(Attendance(
                    date=attendance_date, student=student, batch=batch,
                    organization=org, status=status, marked_by=request.user, notes=notes,
                ))
        with transaction.atomic():
            if to_create:
                Attendance.objects.bulk_create(to_create)
            if to_update:
                Attendance.objects.bulk_update(to_update, ['status', 'marked_by', 'notes'])
        marked_count = len(students)

        messages.success(request, f'Attendance marked for {marked_count} students!')
        return redirect('attendance_list')

    form = AttendanceFilterForm()
    form.fields['batch'].queryset = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    students_data = None
    selected_batch = None
    selected_date = None

    if request.GET.get('batch') and request.GET.get('date'):
        try:
            selected_batch = Batch.objects.select_related('course').get(pk=request.GET['batch'], organization=org)
            selected_date = request.GET['date']
            students = selected_batch.students.filter(organization=org)

            existing = Attendance.objects.filter(
                batch=selected_batch, date=selected_date, organization=org
            )
            existing_attendance = {a.student_id: a for a in existing}

            students_data = []
            for student in students:
                att = existing_attendance.get(student.pk)
                students_data.append({
                    'student': student,
                    'status': att.status if att else 'Present',
                    'notes': att.notes if att else '',
                })

            form.initial = {'batch': selected_batch.pk, 'date': selected_date}
        except Batch.DoesNotExist:
            pass

    context = {
        'form': form,
        'students_data': students_data,
        'selected_batch': selected_batch,
        'selected_date': selected_date,
    }
    return render(request, 'management/attendance_mark.html', context)


@login_required(login_url='login')
@internal_user_required
def quick_attendance(request, batch_id):
    """Quick attendance view - tap to toggle attendance status"""
    org = get_org(request)
    batch = get_object_or_404(Batch.objects.select_related('course'), pk=batch_id, organization=org)
    today = date.today()
    attendance_date = request.GET.get('date', str(today))

    students = batch.students.filter(organization=org).order_by('full_name')

    # Get existing attendance for this date
    existing = Attendance.objects.filter(
        batch=batch, date=attendance_date, organization=org
    )
    attendance_map = {a.student_id: a.status for a in existing}

    # Prepare students with their attendance status
    students_data = []
    for student in students:
        students_data.append({
            'student': student,
            'status': attendance_map.get(student.pk, None),  # None means not marked
        })

    context = {
        'batch': batch,
        'students_data': students_data,
        'attendance_date': attendance_date,
        'today': str(today),
        'present_count': sum(1 for s in students_data if s['status'] == 'Present'),
        'absent_count': sum(1 for s in students_data if s['status'] == 'Absent'),
        'total_count': len(students_data),
    }
    return render(request, 'management/quick_attendance.html', context)


@login_required(login_url='login')
@internal_user_required
@require_POST
def toggle_attendance(request):
    """AJAX endpoint to toggle attendance status"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        batch_id = data.get('batch_id')
        attendance_date = data.get('date')
        new_status = data.get('status')  # Present, Absent, or None (to delete)

        student = get_object_or_404(Student, pk=student_id, organization=org)
        batch = get_object_or_404(Batch, pk=batch_id, organization=org)

        if new_status in ['Present', 'Absent', 'Late', 'Excused']:
            attendance, created = Attendance.objects.update_or_create(
                date=attendance_date,
                student=student,
                batch=batch,
                organization=org,
                defaults={
                    'status': new_status,
                    'marked_by': request.user,
                }
            )
            return JsonResponse({
                'success': True,
                'status': new_status,
                'message': f'{student.full_name} marked as {new_status}'
            })
        else:
            # Delete the attendance record if status is None/empty
            Attendance.objects.filter(
                date=attendance_date,
                student=student,
                batch=batch,
                organization=org
            ).delete()
            return JsonResponse({
                'success': True,
                'status': None,
                'message': f'{student.full_name} attendance cleared'
            })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='login')
@internal_user_required
@require_POST
def mark_all_present(request):
    """AJAX endpoint to mark all students present for a batch on a date"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        batch_id = data.get('batch_id')
        attendance_date = data.get('date')

        batch = get_object_or_404(Batch, pk=batch_id, organization=org)
        students = list(batch.students.filter(organization=org))
        existing = {
            a.student_id: a for a in Attendance.objects.filter(
                date=attendance_date, batch=batch, organization=org,
                student__in=students
            )
        }
        to_create = []
        to_update = []
        for student in students:
            if student.pk in existing:
                att = existing[student.pk]
                att.status = 'Present'
                att.marked_by = request.user
                to_update.append(att)
            else:
                to_create.append(Attendance(
                    date=attendance_date, student=student, batch=batch,
                    organization=org, status='Present', marked_by=request.user,
                ))
        with transaction.atomic():
            if to_create:
                Attendance.objects.bulk_create(to_create)
            if to_update:
                Attendance.objects.bulk_update(to_update, ['status', 'marked_by'])
        count = len(students)

        return JsonResponse({
            'success': True,
            'count': count,
            'message': f'All {count} students marked as Present'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='login')
@internal_user_required
@require_POST
def mark_all_absent(request):
    """AJAX endpoint to mark all students absent for a batch on a date"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        batch_id = data.get('batch_id')
        attendance_date = data.get('date')

        batch = get_object_or_404(Batch, pk=batch_id, organization=org)
        students = list(batch.students.filter(organization=org))
        existing = {
            a.student_id: a for a in Attendance.objects.filter(
                date=attendance_date, batch=batch, organization=org,
                student__in=students
            )
        }
        to_create = []
        to_update = []
        for student in students:
            if student.pk in existing:
                att = existing[student.pk]
                att.status = 'Absent'
                att.marked_by = request.user
                to_update.append(att)
            else:
                to_create.append(Attendance(
                    date=attendance_date, student=student, batch=batch,
                    organization=org, status='Absent', marked_by=request.user,
                ))
        with transaction.atomic():
            if to_create:
                Attendance.objects.bulk_create(to_create)
            if to_update:
                Attendance.objects.bulk_update(to_update, ['status', 'marked_by'])
        count = len(students)

        return JsonResponse({
            'success': True,
            'count': count,
            'message': f'All {count} students marked as Absent'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ─── Staff Attendance Views ───────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def staff_attendance_list(request):
    org = get_org(request)
    attendances = StaffAttendance.objects.filter(organization=org).select_related('staff', 'marked_by')

    filter_date = request.GET.get('date')
    staff_role = request.GET.get('role')
    if filter_date:
        attendances = attendances.filter(date=filter_date)
    if staff_role:
        attendances = attendances.filter(staff__staff_role=staff_role)

    # Summary stats (before pagination)
    summary = attendances.aggregate(
        total_records=Count('id'),
        present_count=Count('id', filter=Q(status='Present')),
        absent_count=Count('id', filter=Q(status='Absent')),
        late_count=Count('id', filter=Q(status='Late')),
        total_hours=Sum('hours'),
    )

    # Calculate total earnings from filtered results using DB aggregation
    total_earnings = 0
    if summary['total_hours']:
        from django.db.models import F
        total_earnings = attendances.filter(
            hours__isnull=False, staff__working_hours_per_day__gt=0
        ).aggregate(
            total=Coalesce(
                Sum(F('hours') * F('staff__salary') / (F('staff__working_hours_per_day') * 26)),
                0, output_field=DecimalField()
            )
        )['total']
        total_earnings = round(float(total_earnings), 2)

    paginator = Paginator(attendances, 50)
    page_number = request.GET.get('page')
    attendances_page = paginator.get_page(page_number)

    context = {
        'attendances': attendances_page,
        'selected_date': filter_date,
        'selected_role': staff_role,
        'role_choices': Staff.ROLE_CHOICES,
        'summary': summary,
        'total_earnings': round(total_earnings, 2),
    }
    return render(request, 'management/staff_attendance_list.html', context)


@login_required(login_url='login')
@internal_user_required
def staff_attendance_mark(request):
    org = get_org(request)
    if request.method == 'POST':
        attendance_date = request.POST.get('date')
        staff_members = list(Staff.objects.filter(organization=org))
        existing = {
            a.staff_id: a for a in StaffAttendance.objects.filter(
                date=attendance_date, organization=org,
                staff__in=staff_members
            )
        }
        to_create = []
        to_update = []
        for staff in staff_members:
            status = request.POST.get(f'status_{staff.pk}', 'Absent')
            notes = request.POST.get(f'notes_{staff.pk}', '')
            hours_str = request.POST.get(f'hours_{staff.pk}', '')
            hours = float(hours_str) if hours_str else None
            if staff.pk in existing:
                att = existing[staff.pk]
                att.status = status
                att.marked_by = request.user
                att.notes = notes
                att.hours = hours
                to_update.append(att)
            else:
                to_create.append(StaffAttendance(
                    date=attendance_date, staff=staff,
                    organization=org, status=status,
                    marked_by=request.user, notes=notes, hours=hours,
                ))
        with transaction.atomic():
            if to_create:
                StaffAttendance.objects.bulk_create(to_create)
            if to_update:
                StaffAttendance.objects.bulk_update(to_update, ['status', 'marked_by', 'notes', 'hours'])
        marked_count = len(staff_members)
        messages.success(request, f'Staff attendance marked for {marked_count} members!')
        return redirect('staff_attendance_list')

    form = StaffAttendanceFilterForm()
    staff_data = None
    selected_date = None

    if request.GET.get('date'):
        selected_date = request.GET['date']
        staff_members = Staff.objects.filter(organization=org).order_by('first_name', 'last_name')
        existing = StaffAttendance.objects.filter(
            date=selected_date, organization=org
        )
        existing_attendance = {a.staff_id: a for a in existing}

        staff_data = []
        for staff in staff_members:
            att = existing_attendance.get(staff.pk)
            hours = att.hours if att else ''
            earnings = round(float(hours) * staff.hourly_rate, 2) if hours else 0
            staff_data.append({
                'staff': staff,
                'status': att.status if att else 'Present',
                'notes': att.notes if att else '',
                'hours': hours,
                'earnings': earnings,
            })

        form.initial = {'date': selected_date}

    context = {
        'form': form,
        'staff_data': staff_data,
        'selected_date': selected_date,
    }
    return render(request, 'management/staff_attendance_mark.html', context)


@login_required(login_url='login')
@internal_user_required
def staff_quick_attendance(request):
    """Quick staff attendance view - tap to toggle attendance status"""
    org = get_org(request)
    today = date.today()
    attendance_date = request.GET.get('date', str(today))

    staff_members = Staff.objects.filter(organization=org).order_by('first_name', 'last_name')

    existing = StaffAttendance.objects.filter(
        date=attendance_date, organization=org
    )
    attendance_map = {a.staff_id: a for a in existing}

    staff_data = []
    for staff in staff_members:
        att = attendance_map.get(staff.pk)
        staff_data.append({
            'staff': staff,
            'status': att.status if att else None,
            'hours': att.hours if att else None,
        })

    context = {
        'staff_data': staff_data,
        'attendance_date': attendance_date,
        'today': str(today),
        'present_count': sum(1 for s in staff_data if s['status'] == 'Present'),
        'absent_count': sum(1 for s in staff_data if s['status'] == 'Absent'),
        'total_count': len(staff_data),
    }
    return render(request, 'management/staff_quick_attendance.html', context)


@login_required(login_url='login')
@internal_user_required
@require_POST
def staff_toggle_attendance(request):
    """AJAX endpoint to toggle staff attendance status"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        staff_id = data.get('staff_id')
        attendance_date = data.get('date')
        new_status = data.get('status')

        staff = get_object_or_404(Staff, pk=staff_id, organization=org)
        hours = data.get('hours')

        if new_status in ['Present', 'Absent', 'Late', 'Excused']:
            defaults = {
                'status': new_status,
                'marked_by': request.user,
            }
            if hours is not None:
                defaults['hours'] = float(hours) if hours else None
            StaffAttendance.objects.update_or_create(
                date=attendance_date,
                staff=staff,
                organization=org,
                defaults=defaults,
            )
            return JsonResponse({
                'success': True,
                'status': new_status,
                'message': f'{staff.first_name} marked as {new_status}'
            })
        else:
            StaffAttendance.objects.filter(
                date=attendance_date,
                staff=staff,
                organization=org
            ).delete()
            return JsonResponse({
                'success': True,
                'status': None,
                'message': f'{staff.first_name} attendance cleared'
            })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='login')
@internal_user_required
@require_POST
def staff_mark_all_present(request):
    """AJAX endpoint to mark all staff present for a date"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        attendance_date = data.get('date')

        staff_members = list(Staff.objects.filter(organization=org))
        existing = {
            a.staff_id: a for a in StaffAttendance.objects.filter(
                date=attendance_date, organization=org,
                staff__in=staff_members
            )
        }
        to_create = []
        to_update = []
        for staff in staff_members:
            if staff.pk in existing:
                att = existing[staff.pk]
                att.status = 'Present'
                att.marked_by = request.user
                to_update.append(att)
            else:
                to_create.append(StaffAttendance(
                    date=attendance_date, staff=staff,
                    organization=org, status='Present',
                    marked_by=request.user,
                ))
        with transaction.atomic():
            if to_create:
                StaffAttendance.objects.bulk_create(to_create)
            if to_update:
                StaffAttendance.objects.bulk_update(to_update, ['status', 'marked_by'])
        count = len(staff_members)

        return JsonResponse({
            'success': True,
            'count': count,
            'message': f'All {count} staff marked as Present'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='login')
@internal_user_required
@require_POST
def staff_mark_all_absent(request):
    """AJAX endpoint to mark all staff absent for a date"""
    org = get_org(request)
    try:
        data = json.loads(request.body)
        attendance_date = data.get('date')

        staff_members = list(Staff.objects.filter(organization=org))
        existing = {
            a.staff_id: a for a in StaffAttendance.objects.filter(
                date=attendance_date, organization=org,
                staff__in=staff_members
            )
        }
        to_create = []
        to_update = []
        for staff in staff_members:
            if staff.pk in existing:
                att = existing[staff.pk]
                att.status = 'Absent'
                att.marked_by = request.user
                to_update.append(att)
            else:
                to_create.append(StaffAttendance(
                    date=attendance_date, staff=staff,
                    organization=org, status='Absent',
                    marked_by=request.user,
                ))
        with transaction.atomic():
            if to_create:
                StaffAttendance.objects.bulk_create(to_create)
            if to_update:
                StaffAttendance.objects.bulk_update(to_update, ['status', 'marked_by'])
        count = len(staff_members)

        return JsonResponse({
            'success': True,
            'count': count,
            'message': f'All {count} staff marked as Absent'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ─── Fee Payment Views ───────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def fee_payment_list(request):
    org = get_org(request)
    payments_qs = FeePayment.objects.filter(organization=org).select_related('student', 'batch__course')

    # Server-side search
    search_query = request.GET.get('q', '').strip()
    if search_query:
        payments_qs = payments_qs.filter(
            Q(receipt_number__icontains=search_query) |
            Q(student__full_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query) |
            Q(batch__course__course_name__icontains=search_query) |
            Q(batch__batch_name__icontains=search_query)
        )

    # Payment method filter
    selected_method = request.GET.get('method', '').strip()
    if selected_method:
        payments_qs = payments_qs.filter(payment_method=selected_method)

    # Status filter
    selected_status = request.GET.get('status', '').strip()
    if selected_status:
        payments_qs = payments_qs.filter(status=selected_status)

    # Count pending for badge
    pending_count = FeePayment.objects.filter(organization=org, status='Pending').count()

    paginator = Paginator(payments_qs, 20)
    page_number = request.GET.get('page')
    payments = paginator.get_page(page_number)
    return render(request, 'management/fee_payment_list.html', {
        'payments': payments,
        'search_query': search_query,
        'selected_method': selected_method,
        'selected_status': selected_status,
        'pending_count': pending_count,
    })


@login_required(login_url='login')
@manager_or_admin_required
def fee_payment_add(request):
    org = get_org(request)
    # Single query for active batches - reused for form dropdown and JS fee data
    active_batches = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    if request.method == 'POST':
        form = FeePaymentForm(request.POST)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        form.fields['batch'].queryset = active_batches
        if form.is_valid():
            payment = form.save(commit=False)
            payment.organization = org
            payment.save()
            messages.success(request, f'Payment recorded! Receipt: {payment.receipt_number}')
            if request.POST.get('submit_action') == 'save_print':
                return redirect('print_receipt', pk=payment.pk)
            return redirect('fee_payment_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FeePaymentForm()
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        form.fields['batch'].queryset = active_batches

        student_uuid = request.GET.get('student')
        if student_uuid:
            try:
                student_obj = Student.objects.get(uuid=student_uuid, organization=org)
                form.initial['student'] = student_obj.pk
            except Student.DoesNotExist:
                pass

    # Build batch fee data for JS auto-calculation (reuse already-fetched batches)
    batch_fees = {}
    for batch in active_batches:
        batch_fees[batch.pk] = {
            'fee': str(batch.course.fees),
            'period': batch.course.fee_period,
            'course_name': f"{batch.course.course_code} - {batch.course.course_name}",
        }

    return render(request, 'management/fee_payment_add.html', {
        'form': form, 'action': 'Record',
        'batch_fees_json': json.dumps(batch_fees),
    })


@login_required(login_url='login')
@manager_or_admin_required
def fee_payment_edit(request, pk):
    org = get_org(request)
    payment = get_object_or_404(FeePayment, pk=pk, organization=org)
    # Single query for active batches - reused for form dropdown and JS fee data
    active_batches = Batch.objects.filter(organization=org, is_active=True).select_related('course').prefetch_related('teachers')

    if request.method == 'POST':
        form = FeePaymentForm(request.POST, instance=payment)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        form.fields['batch'].queryset = active_batches
        if form.is_valid():
            form.save()
            messages.success(request, f'Payment #{payment.receipt_number} updated!')
            if request.POST.get('submit_action') == 'save_print':
                return redirect('print_receipt', pk=payment.pk)
            return redirect('fee_payment_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FeePaymentForm(instance=payment)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        form.fields['batch'].queryset = active_batches

    # Build batch fee data for JS auto-calculation (reuse already-fetched batches)
    batch_fees = {}
    for batch in active_batches:
        batch_fees[batch.pk] = {
            'fee': str(batch.course.fees),
            'period': batch.course.fee_period,
            'course_name': f"{batch.course.course_code} - {batch.course.course_name}",
        }

    return render(request, 'management/fee_payment_add.html', {
        'form': form, 'action': 'Edit',
        'batch_fees_json': json.dumps(batch_fees),
    })


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def fee_payment_delete(request, pk):
    org = get_org(request)
    payment = get_object_or_404(FeePayment, pk=pk, organization=org)
    receipt_number = payment.receipt_number
    payment.delete()
    messages.success(request, f'Payment record #{receipt_number} deleted successfully!')
    return redirect('fee_payment_list')


@login_required(login_url='login')
@internal_user_required
def print_receipt(request, pk):
    org = get_org(request)
    payment = get_object_or_404(FeePayment.objects.select_related('student', 'batch__course'), pk=pk, organization=org)
    student = payment.student
    total_fees = student.get_effective_fee()
    total_paid = student.get_total_paid()
    pending_fees = student.get_pending_fees()

    # Calculate per-month fee for breakdown
    per_month_fee = 0
    months_count = payment.fee_months_count
    if months_count > 0:
        per_month_fee = payment.amount / months_count

    return render(request, 'management/receipt_print.html', {
        'payment': payment,
        'organization': org,
        'total_fees': total_fees,
        'total_paid': total_paid,
        'pending_fees': pending_fees,
        'per_month_fee': per_month_fee,
    })


# ─── Behavior Notes Views ────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def behavior_note_add(request, student_uuid):
    org = get_org(request)
    student = get_object_or_404(Student, uuid=student_uuid, organization=org)
    if request.method == 'POST':
        form = BehaviorNoteForm(request.POST)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        if form.is_valid():
            note = form.save(commit=False)
            note.organization = org
            note.noted_by = request.user
            note.save()
            messages.success(request, f'Behavior note added for {note.student.full_name}.')
            return redirect('student_detail', uuid=note.student.uuid)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BehaviorNoteForm(initial={'student': student.pk})
        form.fields['student'].queryset = Student.objects.filter(organization=org)
    return render(request, 'management/behavior_note_form.html', {'form': form, 'action': 'Add', 'student': student})


@login_required(login_url='login')
@internal_user_required
def behavior_note_edit(request, pk):
    org = get_org(request)
    note = get_object_or_404(BehaviorNote, pk=pk, organization=org)
    if request.method == 'POST':
        form = BehaviorNoteForm(request.POST, instance=note)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
        if form.is_valid():
            form.save()
            messages.success(request, 'Behavior note updated.')
            return redirect('student_detail', uuid=note.student.uuid)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BehaviorNoteForm(instance=note)
        form.fields['student'].queryset = Student.objects.filter(organization=org)
    return render(request, 'management/behavior_note_form.html', {'form': form, 'action': 'Edit', 'student': note.student})


@login_required(login_url='login')
@internal_user_required
@require_POST
def behavior_note_delete(request, pk):
    org = get_org(request)
    note = get_object_or_404(BehaviorNote, pk=pk, organization=org)
    student_uuid = note.student.uuid
    note.delete()
    messages.success(request, 'Behavior note deleted.')
    return redirect('student_detail', uuid=student_uuid)


# ─── API: Cities by State ────────────────────────────────────────────────────

def get_cities_for_state(request):
    """Return cities for a given state as JSON (used by dynamic dropdowns)."""
    state = request.GET.get('state', '')
    cities = sorted(CITY_DATA.get(state, {}).keys())
    return JsonResponse({'cities': cities})


@login_required(login_url='login')
@manager_or_admin_required
def api_student_batches(request):
    """Return a student's enrolled batches with fee info as JSON."""
    org = get_org(request)
    student_id = request.GET.get('student_id', '')
    if not student_id:
        return JsonResponse({'batches': [], 'discount_type': '', 'discount_value': '0'})
    try:
        student = Student.objects.get(pk=student_id, organization=org)
    except Student.DoesNotExist:
        return JsonResponse({'batches': [], 'discount_type': '', 'discount_value': '0'})

    batches = []
    for batch in student.batches.filter(is_active=True).select_related('course'):
        batches.append({
            'id': batch.pk,
            'label': f"{batch.batch_code} - {batch.batch_name} ({batch.course.course_name})",
            'fee': str(batch.course.fees),
            'fee_period': batch.course.fee_period,
            'course_name': f"{batch.course.course_code} - {batch.course.course_name}",
        })

    return JsonResponse({
        'batches': batches,
        'discount_type': student.discount_type or '',
        'discount_value': str(student.discount_value) if student.discount_value > 0 else '0',
    })


# ─── Settings ────────────────────────────────────────────────────────────────

@login_required(login_url='login')
@admin_required
def settings_view(request):
    org = get_org(request)
    if request.method == 'POST':
        form = SettingsForm(request.POST, instance=org)
        if form.is_valid():
            form.save()
            messages.success(request, 'Settings updated successfully!')
            return redirect('settings')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SettingsForm(instance=org)
    return render(request, 'management/settings.html', {'form': form})


# ─── User Management Views ───────────────────────────────────────────────────

@login_required(login_url='login')
@admin_required
def user_list(request):
    """List all users in the organization."""
    org = get_org(request)
    users = User.objects.filter(organization=org).order_by('-date_joined')
    return render(request, 'management/user_list.html', {'users': users})


@login_required(login_url='login')
@admin_required
def user_invite(request):
    """Invite a new user to the organization."""
    org = get_org(request)
    if request.method == 'POST':
        form = InviteUserForm(request.POST, organization=org)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'User "{user.username}" created successfully!')
            return redirect('user_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = InviteUserForm(organization=org)
    return render(request, 'management/user_invite.html', {'form': form})


@login_required(login_url='login')
@admin_required
def user_edit(request, pk):
    """Edit a user's role and details."""
    org = get_org(request)
    user = get_object_or_404(User, pk=pk, organization=org)

    # Prevent editing yourself via this form
    if user == request.user:
        messages.error(request, 'You cannot edit your own account here.')
        return redirect('user_list')

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f'User "{user.username}" updated successfully!')
            return redirect('user_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserEditForm(instance=user)
    return render(request, 'management/user_edit.html', {'form': form, 'edit_user': user})


@login_required(login_url='login')
@admin_required
@require_POST
def user_delete(request, pk):
    """Delete a user from the organization."""
    org = get_org(request)
    user = get_object_or_404(User, pk=pk, organization=org)

    # Prevent deleting yourself
    if user == request.user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('user_list')

    username = user.username
    user.delete()
    messages.success(request, f'User "{username}" deleted successfully!')
    return redirect('user_list')



# ─── Staff Leave Views ──────────────────────────────────────────────────────

@login_required(login_url='login')
@internal_user_required
def staff_leave_list(request):
    """List leave requests. Staff sees own, admin/manager sees all."""
    org = get_org(request)
    leaves = LeaveRequest.objects.filter(organization=org).select_related('staff', 'leave_type', 'reviewed_by')

    # Staff role: only show own leaves
    if request.user.role == 'staff' and request.user.staff_profile:
        leaves = leaves.filter(staff=request.user.staff_profile)

    status_filter = request.GET.get('status', '').strip()
    if status_filter in ('pending', 'approved', 'rejected', 'cancelled'):
        leaves = leaves.filter(status=status_filter)

    search_query = request.GET.get('q', '').strip()
    if search_query:
        leaves = leaves.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(leave_type__name__icontains=search_query)
        )

    status_counts = LeaveRequest.objects.filter(organization=org)
    if request.user.role == 'staff' and request.user.staff_profile:
        status_counts = status_counts.filter(staff=request.user.staff_profile)
    status_counts = status_counts.values('status').annotate(count=Count('id'))
    counts = {item['status']: item['count'] for item in status_counts}

    paginator = Paginator(leaves, 20)
    leaves_page = paginator.get_page(request.GET.get('page'))

    return render(request, 'management/staff_leave_list.html', {
        'leaves': leaves_page,
        'status_filter': status_filter,
        'search_query': search_query,
        'pending_count': counts.get('pending', 0),
        'approved_count': counts.get('approved', 0),
        'rejected_count': counts.get('rejected', 0),
        'cancelled_count': counts.get('cancelled', 0),
        'total_count': sum(counts.values()),
    })


@login_required(login_url='login')
@internal_user_required
def staff_leave_request(request):
    """Request leave. Staff for self, admin/manager for any staff."""
    org = get_org(request)

    # Pre-fill staff from query param (e.g., from staff detail page)
    initial = {}
    staff_id = request.GET.get('staff')
    if staff_id and request.user.can_create_edit():
        try:
            initial['staff'] = Staff.objects.get(pk=staff_id, organization=org)
        except Staff.DoesNotExist:
            pass

    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, organization=org, user=request.user)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.organization = org
            leave.requested_by = request.user
            leave.days = form.cleaned_data.get('_days', 1)
            if leave.half_day:
                leave.end_date = leave.start_date
                leave.days = 0.5
            elif not leave.days:
                leave.days = (leave.end_date - leave.start_date).days + 1
            # Ensure balances exist
            ensure_leave_balances(leave.staff, leave.start_date.year)
            leave.save()
            messages.success(request, f'Leave request submitted for {leave.staff.first_name} {leave.staff.last_name}.')
            return redirect('staff_leave_list')
    else:
        form = LeaveRequestForm(organization=org, user=request.user, initial=initial)

    return render(request, 'management/staff_leave_request.html', {
        'form': form,
    })


@login_required(login_url='login')
@internal_user_required
def staff_leave_detail(request, pk):
    """View leave request details."""
    org = get_org(request)
    leave = get_object_or_404(LeaveRequest, pk=pk, organization=org)

    # Staff can only view their own
    if request.user.role == 'staff' and request.user.staff_profile:
        if leave.staff != request.user.staff_profile:
            messages.error(request, 'You do not have permission to view this leave request.')
            return redirect('staff_leave_list')

    reject_form = LeaveRejectForm()
    return render(request, 'management/staff_leave_detail.html', {
        'leave': leave,
        'reject_form': reject_form,
    })


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def staff_leave_approve(request, pk):
    """Approve a leave request."""
    from django.utils import timezone
    from datetime import timedelta

    org = get_org(request)
    leave = get_object_or_404(LeaveRequest, pk=pk, organization=org)

    if leave.status != 'pending':
        messages.error(request, 'This leave request has already been processed.')
        return redirect('staff_leave_detail', pk=pk)

    with transaction.atomic():
        # Ensure balance exists and deduct
        ensure_leave_balances(leave.staff, leave.start_date.year)
        balance = LeaveBalance.objects.select_for_update().get(
            organization=org, staff=leave.staff,
            leave_type=leave.leave_type, year=leave.start_date.year
        )

        # Check sufficient balance for paid leave
        if leave.leave_type.days_per_year > 0 and balance.remaining < float(leave.days):
            messages.error(request, f'Insufficient leave balance. {balance.remaining} days remaining.')
            return redirect('staff_leave_detail', pk=pk)

        balance.used += leave.days
        balance.save()

        # Update leave status
        leave.status = 'approved'
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.save()

        # Auto-create StaffAttendance as "Excused" for each leave day
        current_date = leave.start_date
        while current_date <= leave.end_date:
            StaffAttendance.objects.update_or_create(
                organization=org, staff=leave.staff, date=current_date,
                defaults={
                    'status': 'Excused',
                    'marked_by': request.user,
                    'notes': f'Leave: {leave.leave_type.name}',
                    'hours': 0,
                }
            )
            current_date += timedelta(days=1)

    messages.success(request, f'Leave approved for {leave.staff.first_name} {leave.staff.last_name}.')
    return redirect('staff_leave_list')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def staff_leave_reject(request, pk):
    """Reject a leave request."""
    from django.utils import timezone

    org = get_org(request)
    leave = get_object_or_404(LeaveRequest, pk=pk, organization=org)

    if leave.status != 'pending':
        messages.error(request, 'This leave request has already been processed.')
        return redirect('staff_leave_detail', pk=pk)

    form = LeaveRejectForm(request.POST)
    if form.is_valid():
        leave.status = 'rejected'
        leave.rejection_reason = form.cleaned_data.get('rejection_reason', '')
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.save()
        messages.success(request, f'Leave request from {leave.staff.first_name} {leave.staff.last_name} rejected.')
    return redirect('staff_leave_list')


@login_required(login_url='login')
@internal_user_required
@require_POST
def staff_leave_cancel(request, pk):
    """Cancel a leave request. Staff cancels own pending. Manager can cancel any."""
    from django.utils import timezone
    from datetime import timedelta

    org = get_org(request)
    leave = get_object_or_404(LeaveRequest, pk=pk, organization=org)

    # Staff can only cancel their own pending leaves
    if request.user.role == 'staff':
        if not request.user.staff_profile or leave.staff != request.user.staff_profile:
            messages.error(request, 'You can only cancel your own leave requests.')
            return redirect('staff_leave_list')
        if leave.status != 'pending':
            messages.error(request, 'You can only cancel pending leave requests.')
            return redirect('staff_leave_detail', pk=pk)

    if leave.status not in ('pending', 'approved'):
        messages.error(request, 'This leave request cannot be cancelled.')
        return redirect('staff_leave_detail', pk=pk)

    with transaction.atomic():
        # If it was approved, restore balance and clean up attendance
        if leave.status == 'approved':
            balance = LeaveBalance.objects.filter(
                organization=org, staff=leave.staff,
                leave_type=leave.leave_type, year=leave.start_date.year
            ).first()
            if balance:
                balance.used = max(0, float(balance.used) - float(leave.days))
                balance.save()

            # Delete auto-created attendance records
            current_date = leave.start_date
            while current_date <= leave.end_date:
                StaffAttendance.objects.filter(
                    organization=org, staff=leave.staff, date=current_date,
                    notes__startswith='Leave:'
                ).delete()
                current_date += timedelta(days=1)

        leave.status = 'cancelled'
        leave.reviewed_by = request.user
        leave.reviewed_at = timezone.now()
        leave.save()

    messages.success(request, f'Leave request cancelled.')
    return redirect('staff_leave_list')


# ─── Admission Application Views ────────────────────────────────────────────

def admission_apply(request, org_slug):
    """Public admission application form. No login required."""
    org = get_object_or_404(Organization, slug=org_slug)

    if request.method == 'POST':
        form = AdmissionApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.organization = org
            application.save()
            return redirect('admission_apply_success', org_slug=org_slug)
    else:
        form = AdmissionApplicationForm()

    return render(request, 'management/admission_apply.html', {
        'form': form,
        'organization': org,
    })


def admission_apply_success(request, org_slug):
    """Success page after submitting admission application."""
    org = get_object_or_404(Organization, slug=org_slug)
    return render(request, 'management/admission_apply_success.html', {
        'organization': org,
    })


@login_required(login_url='login')
@manager_or_admin_required
def application_list(request):
    """Redirect to student list with applications tab."""
    url = reverse('student_list') + '?tab=applications'
    status = request.GET.get('status', '')
    if status:
        url += f'&status={status}'
    q = request.GET.get('q', '')
    if q:
        url += f'&q={q}'
    return redirect(url)


def _application_list_legacy(request):
    """Legacy: kept for reference. No longer used directly."""
    org = get_org(request)
    applications = AdmissionApplication.objects.filter(organization=org)

    status_filter = request.GET.get('status', '').strip()
    if status_filter in ('pending', 'accepted', 'rejected'):
        applications = applications.filter(status=status_filter)

    search_query = request.GET.get('q', '').strip()
    if search_query:
        applications = applications.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    status_counts = AdmissionApplication.objects.filter(organization=org).values('status').annotate(
        count=Count('id')
    )
    counts = {item['status']: item['count'] for item in status_counts}

    paginator = Paginator(applications, 20)
    applications_page = paginator.get_page(request.GET.get('page'))

    return render(request, 'management/application_list.html', {
        'applications': applications_page,
        'status_filter': status_filter,
        'search_query': search_query,
        'pending_count': counts.get('pending', 0),
        'accepted_count': counts.get('accepted', 0),
        'rejected_count': counts.get('rejected', 0),
        'total_count': sum(counts.values()),
    })


@login_required(login_url='login')
@manager_or_admin_required
def application_detail(request, pk):
    """View admission application details."""
    org = get_org(request)
    application = get_object_or_404(AdmissionApplication, pk=pk, organization=org)
    reject_form = ApplicationRejectForm()
    return render(request, 'management/application_detail.html', {
        'application': application,
        'reject_form': reject_form,
    })


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def application_accept(request, pk):
    """Accept an admission application and create a Student record."""
    from django.utils import timezone

    org = get_org(request)
    application = get_object_or_404(AdmissionApplication, pk=pk, organization=org)

    if application.status != 'pending':
        messages.error(request, 'This application has already been processed.')
        return redirect('application_detail', pk=pk)

    with transaction.atomic():
        student = Student(
            full_name=f"{application.first_name} {application.last_name}".strip(),
            phone=application.phone,
            email=application.email,
            date_of_birth=application.date_of_birth,
            gender=application.gender,
            address=application.address,
            city=application.city,
            state=application.state,
            pin_code=application.pin_code,
            photo=application.photo if application.photo else None,
            enrollment_date=date.today(),
            organization=org,
        )
        student.save()

        application.status = 'accepted'
        application.student = student
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.save()

    messages.success(
        request,
        f'Application accepted! Student "{student.full_name}" '
        f'created with ID {student.student_id}. You can now assign batches.'
    )
    return redirect('student_edit', uuid=student.uuid)


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def application_reject(request, pk):
    """Reject an admission application."""
    from django.utils import timezone

    org = get_org(request)
    application = get_object_or_404(AdmissionApplication, pk=pk, organization=org)

    if application.status != 'pending':
        messages.error(request, 'This application has already been processed.')
        return redirect('application_detail', pk=pk)

    form = ApplicationRejectForm(request.POST)
    if form.is_valid():
        application.status = 'rejected'
        application.rejection_reason = form.cleaned_data.get('rejection_reason', '')
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.save()

        messages.success(
            request,
            f'Application from "{application.first_name} {application.last_name}" has been rejected.'
        )
    return redirect('application_list')


# ─── Parent Portal Views ────────────────────────────────────────────────────

@login_required(login_url='login')
@parent_required
def parent_dashboard(request):
    """Parent portal: show all students linked to this parent's phone number."""
    from django.db.models import Prefetch
    from .utils import normalize_phone

    user = request.user
    org = user.organization

    # Extract phone from org-scoped username (format: phone_orgId)
    parent_phone = user.username.rsplit('_', 1)[0] if '_' in user.username else user.username

    # First pass: find matching student IDs (phones stored raw, username contains normalized phone)
    all_students = Student.objects.filter(organization=org).only('id', 'phone')
    matched_ids = [s.id for s in all_students if normalize_phone(s.phone) == parent_phone]

    # Second pass: load matched students with all related data in bulk
    matched_students = Student.objects.filter(
        id__in=matched_ids
    ).prefetch_related(
        'batches__course',
        Prefetch(
            'attendances',
            queryset=Attendance.objects.select_related('batch__course').order_by('-date'),
            to_attr='all_attendances'
        ),
        Prefetch(
            'fee_payments',
            queryset=FeePayment.objects.select_related('batch__course').order_by('-payment_date'),
            to_attr='all_fee_payments'
        ),
        Prefetch(
            'behavior_notes',
            queryset=BehaviorNote.objects.select_related('noted_by').order_by('-date'),
            to_attr='all_behavior_notes'
        ),
    ).annotate(
        att_total=Count('attendances'),
        att_present=Count('attendances', filter=Q(attendances__status='Present')),
        att_absent=Count('attendances', filter=Q(attendances__status='Absent')),
        att_late=Count('attendances', filter=Q(attendances__status='Late')),
        total_fees=Sum('batches__course__fees'),
        total_paid=Coalesce(Sum('fee_payments__amount', filter=Q(fee_payments__status='Approved')), 0, output_field=DecimalField()),
    )

    students_data = []
    for student in matched_students:
        total_fees = student.total_fees or 0
        total_paid = student.total_paid or 0
        att_total = student.att_total or 0
        att_present = student.att_present or 0

        if att_total > 0:
            attendance_pct = round(((att_present + (student.att_late or 0)) / att_total) * 100, 1)
        else:
            attendance_pct = 0

        days_enrolled = (date.today() - student.enrollment_date).days if student.enrollment_date else 0

        students_data.append({
            'student': student,
            'attendances': student.all_attendances[:20],
            'fee_payments': student.all_fee_payments,
            'behavior_notes': student.all_behavior_notes,
            'attendance_percentage': attendance_pct,
            'total_fees': total_fees,
            'total_paid': total_paid,
            'pending_fees': total_fees - total_paid,
            'att_present': att_present,
            'att_absent': student.att_absent or 0,
            'att_late': student.att_late or 0,
            'att_total': att_total,
            'days_enrolled': days_enrolled,
        })

    # Check if parent is still using the default password (phone number)
    is_default_password = user.check_password(parent_phone)

    context = {
        'students_data': students_data,
        'student_count': len(students_data),
        'is_default_password': is_default_password,
        'org': org,
    }
    return render(request, 'management/parent_dashboard.html', context)


@login_required(login_url='login')
@parent_required
def parent_change_password(request):
    """Allow parent users to change their password."""
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('parent_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'management/parent_change_password.html', {'form': form})


@login_required(login_url='login')
@parent_required
def parent_pay_upi(request):
    """Parent UPI payment: select course/month, open UPI app or show QR code."""
    from urllib.parse import quote
    from .utils import normalize_phone

    user = request.user
    org = user.organization

    if not org or not org.upi_id:
        messages.error(request, 'UPI payments are not configured for this organization.')
        return redirect('parent_dashboard')

    # Find students linked to this parent (same logic as parent_dashboard)
    parent_phone = user.username.rsplit('_', 1)[0] if '_' in user.username else user.username
    all_students = Student.objects.filter(organization=org).only('id', 'phone')
    matched_ids = [s.id for s in all_students if normalize_phone(s.phone) == parent_phone]
    students = Student.objects.filter(id__in=matched_ids).prefetch_related('batches__course')

    # Collect active batches for all linked students
    student_batches = []
    for student in students:
        for batch in student.batches.filter(is_active=True).select_related('course'):
            student_batches.append({
                'student': student,
                'batch': batch,
                'course': batch.course,
                'fees': batch.course.fees,
                'fee_period': batch.course.get_fee_period_display(),
            })

    # Generate month options (current + next 5 months for advance payment)
    today = date.today()
    month_options = []
    for i in range(6):
        m = today.month + i
        y = today.year
        if m > 12:
            m -= 12
            y += 1
        month_options.append({
            'value': f"{y}-{m:02d}",
            'label': date(y, m, 1).strftime('%B %Y'),
        })

    # Check if a specific batch/student/months were selected
    batch_id = request.GET.get('batch_id')
    student_uuid = request.GET.get('student')
    months = request.GET.getlist('month')

    upi_data = None

    if batch_id and student_uuid and months:
        # Validate all months are from our allowed options
        valid_values = {opt['value'] for opt in month_options}
        months = [m for m in months if m in valid_values]
        if not months:
            messages.error(request, 'Invalid month selection. Please try again.')
            return redirect('parent_pay_upi')

        try:
            selected_student = get_object_or_404(Student, uuid=student_uuid, id__in=matched_ids)
            selected_batch = get_object_or_404(
                Batch.objects.select_related('course'),
                pk=batch_id, organization=org, is_active=True
            )
            if not selected_student.batches.filter(pk=selected_batch.pk).exists():
                messages.error(request, 'Student is not enrolled in this batch.')
                return redirect('parent_pay_upi')

            per_month_fee = selected_batch.course.fees
            num_months = len(months)
            total_amount = per_month_fee * num_months

            # Build month labels for display
            month_labels = []
            for m in months:
                for opt in month_options:
                    if opt['value'] == m:
                        month_labels.append(opt['label'])
                        break

            months_display = ', '.join(month_labels)
            months_short = ', '.join(months)
            tn = f"{selected_student.student_id} {selected_student.full_name} - {selected_batch.course.course_name} - {months_short}"

            # UPI apps expect minimal encoding: preserve @, -, . in values
            upi_link = (
                f"upi://pay?"
                f"pa={quote(org.upi_id, safe='@.-')}"
                f"&pn={quote(org.org_name, safe=' ')}"
                f"&am={total_amount}"
                f"&cu=INR"
                f"&tn={quote(tn, safe=' -')}"
            )

            upi_data = {
                'upi_link': upi_link,
                'upi_id': org.upi_id,
                'org_name': org.org_name,
                'per_month_fee': per_month_fee,
                'num_months': num_months,
                'total_amount': total_amount,
                'fee_period': selected_batch.course.get_fee_period_display(),
                'course_name': selected_batch.course.course_name,
                'batch_name': selected_batch.batch_name,
                'student_name': selected_student.full_name,
                'student_id': selected_student.student_id,
                'student_pk': selected_student.pk,
                'batch_pk': selected_batch.pk,
                'months': months_display,
                'months_raw': months,
                'transaction_note': tn,
            }
        except Exception:
            messages.error(request, 'Invalid selection. Please try again.')
            return redirect('parent_pay_upi')

    context = {
        'student_batches': student_batches,
        'upi_data': upi_data,
        'month_options': month_options,
    }
    return render(request, 'management/parent_pay_upi.html', context)


@login_required(login_url='login')
@parent_required
@require_POST
def parent_confirm_payment(request):
    """Parent confirms they have paid via UPI – creates a pending FeePayment."""
    from .utils import normalize_phone

    user = request.user
    org = user.organization

    student_pk = request.POST.get('student_pk')
    batch_pk = request.POST.get('batch_pk')
    amount = request.POST.get('amount')
    months_raw = request.POST.getlist('months_raw')

    if not (student_pk and batch_pk and amount and months_raw):
        messages.error(request, 'Missing payment details. Please try again.')
        return redirect('parent_pay_upi')

    # Verify the student belongs to this parent
    parent_phone = user.username.rsplit('_', 1)[0] if '_' in user.username else user.username
    all_students = Student.objects.filter(organization=org).only('id', 'phone')
    matched_ids = [s.id for s in all_students if normalize_phone(s.phone) == parent_phone]

    try:
        student = Student.objects.get(pk=student_pk, id__in=matched_ids)
        batch = Batch.objects.get(pk=batch_pk, organization=org)
    except (Student.DoesNotExist, Batch.DoesNotExist):
        messages.error(request, 'Invalid student or batch.')
        return redirect('parent_pay_upi')

    # Build fee_month_from / fee_month_to from months_raw (e.g. ['2026-03', '2026-04'])
    months_sorted = sorted(months_raw)
    fee_month_from = date(int(months_sorted[0].split('-')[0]), int(months_sorted[0].split('-')[1]), 1)
    fee_month_to = date(int(months_sorted[-1].split('-')[0]), int(months_sorted[-1].split('-')[1]), 1)

    payment = FeePayment(
        student=student,
        batch=batch,
        amount=amount,
        fee_month_from=fee_month_from,
        fee_month_to=fee_month_to,
        payment_date=date.today(),
        payment_method='UPI',
        status='Pending',
        notes=f'Online payment by parent. Months: {", ".join(months_sorted)}',
        organization=org,
    )
    payment.save()
    messages.success(request, 'Payment submitted! It will be visible once the admin approves it.')
    return redirect('parent_dashboard')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def fee_payment_approve(request, pk):
    org = get_org(request)
    payment = get_object_or_404(FeePayment, pk=pk, organization=org)
    payment.status = 'Approved'
    payment.save()
    messages.success(request, f'Payment #{payment.receipt_number} approved!')
    return redirect('fee_payment_list')


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def fee_payment_reject(request, pk):
    org = get_org(request)
    payment = get_object_or_404(FeePayment, pk=pk, organization=org)
    payment.status = 'Rejected'
    payment.save()
    messages.success(request, f'Payment #{payment.receipt_number} rejected.')
    return redirect('fee_payment_list')


# ─── Calendar / Event Views ──────────────────────────────────────────────────

@login_required(login_url='login')
def calendar_view(request):
    org = get_org(request)
    today = date.today()

    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if month < 1 or month > 12:
            month = today.month
        if year < 2000 or year > 2100:
            year = today.year
    except (ValueError, TypeError):
        year = today.year
        month = today.month

    cal = cal_module.Calendar(firstweekday=6)  # Sunday first
    month_days = cal.monthdayscalendar(year, month)

    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1)
    else:
        last_day = date(year, month + 1, 1)

    events = Event.objects.filter(
        organization=org,
        start_date__lt=last_day,
        end_date__gte=first_day,
    ).order_by('start_date', 'title')

    events_by_day = {}
    for event in events:
        event_start = max(event.start_date, first_day)
        event_end = min(event.end_date, last_day - timedelta(days=1))
        current = event_start
        while current <= event_end:
            if current.month == month and current.year == year:
                events_by_day.setdefault(current.day, []).append(event)
            current += timedelta(days=1)

    weeks = []
    for week in month_days:
        week_data = []
        for day_num in week:
            if day_num == 0:
                week_data.append({'day': 0, 'events': [], 'is_today': False})
            else:
                is_today = (day_num == today.day and month == today.month and year == today.year)
                week_data.append({
                    'day': day_num,
                    'events': events_by_day.get(day_num, []),
                    'is_today': is_today,
                })
        weeks.append(week_data)

    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    month_name = cal_module.month_name[month]

    selected_day = request.GET.get('day')
    selected_day_events = []
    selected_date = None
    if selected_day:
        try:
            selected_day = int(selected_day)
            if 1 <= selected_day <= cal_module.monthrange(year, month)[1]:
                selected_day_events = events_by_day.get(selected_day, [])
                selected_date = date(year, month, selected_day)
            else:
                selected_day = None
        except (ValueError, TypeError):
            selected_day = None

    context = {
        'weeks': weeks,
        'month': month,
        'year': year,
        'month_name': month_name,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'today': today,
        'selected_day': selected_day,
        'selected_day_events': selected_day_events,
        'selected_date': selected_date,
        'can_manage_events': request.user.can_create_edit(),
    }
    return render(request, 'management/calendar.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def event_add(request):
    org = get_org(request)
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.organization = org
            event.created_by = request.user
            event.save()
            messages.success(request, f'Event "{event.title}" created successfully!')
            return redirect(f"{reverse('calendar')}?year={event.start_date.year}&month={event.start_date.month}&day={event.start_date.day}")
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        initial = {}
        day = request.GET.get('day')
        month = request.GET.get('month')
        year = request.GET.get('year')
        if day and month and year:
            try:
                initial['start_date'] = date(int(year), int(month), int(day))
            except (ValueError, TypeError):
                pass
        form = EventForm(initial=initial)
    return render(request, 'management/event_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def event_edit(request, pk):
    org = get_org(request)
    event = get_object_or_404(Event, pk=pk, organization=org)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, f'Event "{event.title}" updated successfully!')
            return redirect(f"{reverse('calendar')}?year={event.start_date.year}&month={event.start_date.month}&day={event.start_date.day}")
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EventForm(instance=event)
    return render(request, 'management/event_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def event_delete(request, pk):
    org = get_org(request)
    event = get_object_or_404(Event, pk=pk, organization=org)
    event_title = event.title
    start = event.start_date
    event.delete()
    messages.success(request, f'Event "{event_title}" deleted successfully!')
    return redirect(f"{reverse('calendar')}?year={start.year}&month={start.month}")


# ─── Staff Self-Service Portal Views ───────────────────────────────────────────

def _calculate_hours_from_punches(punches):
    """Calculate total hours from a queryset of PunchRecords for a single day."""
    total_seconds = 0
    punch_in_time = None
    for punch in punches:
        if punch.punch_type == 'in':
            punch_in_time = punch.timestamp
        elif punch.punch_type == 'out' and punch_in_time:
            total_seconds += (punch.timestamp - punch_in_time).total_seconds()
            punch_in_time = None
    # If still punched in, count time until now
    if punch_in_time:
        from django.utils import timezone
        total_seconds += (timezone.now() - punch_in_time).total_seconds()
    return round(total_seconds / 3600, 1)


@login_required(login_url='login')
@staff_role_required
def staff_portal(request):
    """Staff self-service dashboard."""
    user = request.user
    staff = user.staff_profile
    org = user.organization
    today = date.today()

    # Today's punch status
    today_punches = PunchRecord.objects.filter(
        staff=staff, date=today, organization=org
    ).order_by('timestamp')
    last_punch = today_punches.last()
    is_punched_in = last_punch and last_punch.punch_type == 'in'

    # Calculate hours worked today
    hours_today = _calculate_hours_from_punches(today_punches)

    # This month's attendance summary
    first_of_month = today.replace(day=1)
    month_attendance = StaffAttendance.objects.filter(
        staff=staff, organization=org,
        date__gte=first_of_month, date__lte=today
    ).aggregate(
        total=Count('id'),
        present=Count('id', filter=Q(status='Present')),
        absent=Count('id', filter=Q(status='Absent')),
        late=Count('id', filter=Q(status='Late')),
        total_hours=Coalesce(Sum('hours'), 0, output_field=DecimalField()),
    )

    # Leave balance for current year
    ensure_leave_balances(staff, today.year)
    leave_balances = LeaveBalance.objects.filter(
        staff=staff, year=today.year, organization=org
    ).select_related('leave_type')

    # Pending leave requests
    pending_leaves = LeaveRequest.objects.filter(
        staff=staff, organization=org, status='pending'
    ).count()

    # Latest payroll
    latest_payroll = Payroll.objects.filter(
        staff=staff, organization=org, status__in=['processed', 'paid']
    ).order_by('-year', '-month').first()

    context = {
        'staff': staff,
        'is_punched_in': is_punched_in,
        'last_punch': last_punch,
        'hours_today': hours_today,
        'today_punches': today_punches,
        'month_attendance': month_attendance,
        'leave_balances': leave_balances,
        'pending_leaves': pending_leaves,
        'latest_payroll': latest_payroll,
        'today': today,
    }
    return render(request, 'management/staff_portal.html', context)


@login_required(login_url='login')
@staff_role_required
@require_POST
def staff_punch(request):
    """Toggle punch in/out for staff."""
    staff = request.user.staff_profile
    org = request.user.organization
    today = date.today()

    last_punch = PunchRecord.objects.filter(
        staff=staff, date=today, organization=org
    ).order_by('timestamp').last()

    if last_punch and last_punch.punch_type == 'in':
        punch_type = 'out'
        msg = 'Punched out successfully!'
    else:
        punch_type = 'in'
        msg = 'Punched in successfully!'

    PunchRecord.objects.create(
        staff=staff,
        punch_type=punch_type,
        date=today,
        organization=org,
    )
    messages.success(request, msg)
    return redirect('staff_portal')


@login_required(login_url='login')
@staff_role_required
def staff_my_attendance(request):
    """Staff views their own attendance history."""
    staff = request.user.staff_profile
    org = request.user.organization
    today = date.today()

    attendances = StaffAttendance.objects.filter(
        staff=staff, organization=org
    ).order_by('-date')

    # Filter by month/year
    month = request.GET.get('month', str(today.month))
    year = request.GET.get('year', str(today.year))
    if month and year:
        try:
            month = int(month)
            year = int(year)
            attendances = attendances.filter(date__month=month, date__year=year)
        except (ValueError, TypeError):
            month = today.month
            year = today.year

    summary = attendances.aggregate(
        total=Count('id'),
        present=Count('id', filter=Q(status='Present')),
        absent=Count('id', filter=Q(status='Absent')),
        late=Count('id', filter=Q(status='Late')),
        excused=Count('id', filter=Q(status='Excused')),
        total_hours=Coalesce(Sum('hours'), 0, output_field=DecimalField()),
    )

    paginator = Paginator(attendances, 31)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'attendances': page,
        'summary': summary,
        'selected_month': month,
        'selected_year': year,
        'months': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'years': list(range(today.year - 2, today.year + 1)),
    }
    return render(request, 'management/staff_my_attendance.html', context)


@login_required(login_url='login')
@staff_role_required
def staff_my_profile(request):
    """Staff views their own profile (read-only)."""
    staff = request.user.staff_profile
    org = request.user.organization
    today = date.today()

    ensure_leave_balances(staff, today.year)
    leave_balances = LeaveBalance.objects.filter(
        staff=staff, year=today.year, organization=org
    ).select_related('leave_type')

    context = {
        'staff': staff,
        'leave_balances': leave_balances,
    }
    return render(request, 'management/staff_my_profile.html', context)


@login_required(login_url='login')
@staff_role_required
def staff_my_salary(request):
    """Staff views their salary/payroll history."""
    staff = request.user.staff_profile
    org = request.user.organization

    payrolls = Payroll.objects.filter(
        staff=staff, organization=org, status__in=['processed', 'paid']
    ).order_by('-year', '-month')

    context = {
        'payrolls': payrolls,
        'staff': staff,
    }
    return render(request, 'management/staff_my_salary.html', context)


@login_required(login_url='login')
@staff_role_required
def staff_my_payslip(request, pk):
    """Staff views a single pay slip."""
    staff = request.user.staff_profile
    payroll = get_object_or_404(Payroll, pk=pk, staff=staff, organization=request.user.organization)

    components = payroll.components.all()
    earnings = components.filter(component_type='earning')
    deductions = components.filter(component_type='deduction')

    context = {
        'payroll': payroll,
        'earnings': earnings,
        'deductions': deductions,
        'staff': staff,
        'org': request.user.organization,
    }
    return render(request, 'management/staff_my_payslip.html', context)


@login_required(login_url='login')
@staff_role_required
def staff_change_password(request):
    """Allow staff to change their password."""
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('staff_portal')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'management/staff_change_password.html', {'form': form})


# ─── Leave Type Management Views ──────────────────────────────────────────────

@login_required(login_url='login')
@manager_or_admin_required
def leave_type_list(request):
    """List all leave types for the organization."""
    org = get_org(request)
    leave_types = LeaveType.objects.filter(organization=org)
    return render(request, 'management/leave_type_list.html', {'leave_types': leave_types})


@login_required(login_url='login')
@manager_or_admin_required
def leave_type_add(request):
    """Add a new leave type."""
    org = get_org(request)
    if request.method == 'POST':
        form = LeaveTypeForm(request.POST)
        if form.is_valid():
            leave_type = form.save(commit=False)
            leave_type.organization = org
            leave_type.save()
            messages.success(request, f'Leave type "{leave_type.name}" created successfully!')
            return redirect('leave_type_list')
    else:
        form = LeaveTypeForm()
    return render(request, 'management/leave_type_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def leave_type_edit(request, pk):
    """Edit a leave type."""
    org = get_org(request)
    leave_type = get_object_or_404(LeaveType, pk=pk, organization=org)
    if request.method == 'POST':
        form = LeaveTypeForm(request.POST, instance=leave_type)
        if form.is_valid():
            form.save()
            messages.success(request, f'Leave type "{leave_type.name}" updated successfully!')
            return redirect('leave_type_list')
    else:
        form = LeaveTypeForm(instance=leave_type)
    return render(request, 'management/leave_type_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def leave_type_delete(request, pk):
    """Delete a leave type."""
    org = get_org(request)
    leave_type = get_object_or_404(LeaveType, pk=pk, organization=org)
    name = leave_type.name
    # Check if there are any leave requests using this type
    if LeaveRequest.objects.filter(leave_type=leave_type).exists():
        messages.error(request, f'Cannot delete "{name}" — it has leave requests associated with it.')
        return redirect('leave_type_list')
    leave_type.delete()
    messages.success(request, f'Leave type "{name}" deleted successfully!')
    return redirect('leave_type_list')


# ─── Admin Payroll Management Views ────────────────────────────────────────────

@login_required(login_url='login')
@manager_or_admin_required
def salary_component_list(request):
    """List all salary components for the organization."""
    org = get_org(request)
    components = SalaryComponent.objects.filter(organization=org)
    return render(request, 'management/salary_component_list.html', {'components': components})


@login_required(login_url='login')
@manager_or_admin_required
def salary_component_add(request):
    """Add a new salary component."""
    org = get_org(request)
    if request.method == 'POST':
        form = SalaryComponentForm(request.POST)
        if form.is_valid():
            component = form.save(commit=False)
            component.organization = org
            component.save()
            messages.success(request, f'Salary component "{component.name}" created successfully!')
            return redirect('salary_component_list')
    else:
        form = SalaryComponentForm()
    return render(request, 'management/salary_component_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def salary_component_edit(request, pk):
    """Edit a salary component."""
    org = get_org(request)
    component = get_object_or_404(SalaryComponent, pk=pk, organization=org)
    if request.method == 'POST':
        form = SalaryComponentForm(request.POST, instance=component)
        if form.is_valid():
            form.save()
            messages.success(request, f'Salary component "{component.name}" updated successfully!')
            return redirect('salary_component_list')
    else:
        form = SalaryComponentForm(instance=component)
    return render(request, 'management/salary_component_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def salary_component_delete(request, pk):
    """Delete a salary component."""
    org = get_org(request)
    component = get_object_or_404(SalaryComponent, pk=pk, organization=org)
    name = component.name
    component.delete()
    messages.success(request, f'Salary component "{name}" deleted successfully!')
    return redirect('salary_component_list')


@login_required(login_url='login')
@manager_or_admin_required
def payroll_list(request):
    """List all payrolls with filters."""
    org = get_org(request)
    today = date.today()

    payrolls = Payroll.objects.filter(organization=org).select_related('staff')

    # Filters
    month = request.GET.get('month', str(today.month))
    year = request.GET.get('year', str(today.year))
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '').strip()

    if month and year:
        try:
            payrolls = payrolls.filter(month=int(month), year=int(year))
        except (ValueError, TypeError):
            pass

    if status_filter in ('draft', 'processed', 'paid'):
        payrolls = payrolls.filter(status=status_filter)

    if search_query:
        payrolls = payrolls.filter(
            Q(staff__first_name__icontains=search_query) |
            Q(staff__last_name__icontains=search_query) |
            Q(payroll_number__icontains=search_query)
        )

    # Status counts for current filter
    base_qs = Payroll.objects.filter(organization=org)
    if month and year:
        try:
            base_qs = base_qs.filter(month=int(month), year=int(year))
        except (ValueError, TypeError):
            pass
    status_counts = base_qs.values('status').annotate(count=Count('id'))
    counts = {item['status']: item['count'] for item in status_counts}

    paginator = Paginator(payrolls, 20)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'payrolls': page,
        'selected_month': month,
        'selected_year': year,
        'status_filter': status_filter,
        'search_query': search_query,
        'months': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'years': list(range(today.year - 2, today.year + 1)),
        'draft_count': counts.get('draft', 0),
        'processed_count': counts.get('processed', 0),
        'paid_count': counts.get('paid', 0),
        'total_count': sum(counts.values()),
    }
    return render(request, 'management/payroll_list.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def payroll_generate(request):
    """Generate payroll for all staff for a given month/year."""
    org = get_org(request)
    today = date.today()

    if request.method == 'POST':
        try:
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
        except (TypeError, ValueError):
            messages.error(request, 'Invalid month or year.')
            return redirect('payroll_generate')

        staff_members = Staff.objects.filter(organization=org)
        created_count = 0
        skipped_count = 0

        for staff_member in staff_members:
            if Payroll.objects.filter(organization=org, staff=staff_member, month=month, year=year).exists():
                skipped_count += 1
                continue

            # Calculate attendance for the month
            attendance = StaffAttendance.objects.filter(
                staff=staff_member, organization=org,
                date__month=month, date__year=year
            ).aggregate(
                present=Count('id', filter=Q(status='Present')),
                absent=Count('id', filter=Q(status='Absent')),
                late=Count('id', filter=Q(status='Late')),
                total_hours=Coalesce(Sum('hours'), 0, output_field=DecimalField()),
            )

            payroll = Payroll(
                staff=staff_member,
                month=month,
                year=year,
                base_salary=staff_member.salary,
                total_earnings=staff_member.salary,
                total_deductions=0,
                net_salary=staff_member.salary,
                days_present=attendance['present'] or 0,
                days_absent=attendance['absent'] or 0,
                days_late=attendance['late'] or 0,
                total_hours=attendance['total_hours'] or 0,
                status='draft',
                generated_by=request.user,
                organization=org,
            )
            payroll.save()
            created_count += 1

        msg = f'Generated {created_count} payroll records for {cal_module.month_name[month]} {year}.'
        if skipped_count:
            msg += f' ({skipped_count} already existed.)'
        messages.success(request, msg)
        return redirect(f"{reverse('payroll_list')}?month={month}&year={year}")

    context = {
        'current_month': today.month,
        'current_year': today.year,
        'months': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'years': list(range(today.year - 2, today.year + 1)),
    }
    return render(request, 'management/payroll_generate.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def payroll_detail(request, pk):
    """View a single payroll with components."""
    org = get_org(request)
    payroll = get_object_or_404(Payroll, pk=pk, organization=org)

    components = payroll.components.all()
    earnings = components.filter(component_type='earning')
    deductions = components.filter(component_type='deduction')

    context = {
        'payroll': payroll,
        'earnings': earnings,
        'deductions': deductions,
    }
    return render(request, 'management/payroll_detail.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def payroll_edit(request, pk):
    """Add/remove components on a draft payroll."""
    org = get_org(request)
    payroll = get_object_or_404(Payroll, pk=pk, organization=org)

    if payroll.status != 'draft':
        messages.error(request, 'Only draft payrolls can be edited.')
        return redirect('payroll_detail', pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_component':
            form = PayrollComponentForm(request.POST, organization=org)
            if form.is_valid():
                component = form.save(commit=False)
                component.payroll = payroll
                # If a salary_component was selected, snapshot its name and type
                if component.salary_component:
                    if not component.name:
                        component.name = component.salary_component.name
                    if not component.component_type:
                        component.component_type = component.salary_component.component_type
                component.save()
                payroll.recalculate_totals()
                messages.success(request, f'Component "{component.name}" added.')
            else:
                messages.error(request, 'Please correct the errors below.')

        elif action == 'remove_component':
            component_id = request.POST.get('component_id')
            PayrollComponent.objects.filter(pk=component_id, payroll=payroll).delete()
            payroll.recalculate_totals()
            messages.success(request, 'Component removed.')

        return redirect('payroll_edit', pk=pk)

    form = PayrollComponentForm(organization=org)
    components = payroll.components.all()
    earnings = components.filter(component_type='earning')
    deductions = components.filter(component_type='deduction')

    context = {
        'payroll': payroll,
        'form': form,
        'earnings': earnings,
        'deductions': deductions,
    }
    return render(request, 'management/payroll_edit.html', context)


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def payroll_process(request, pk):
    """Mark a draft payroll as processed."""
    org = get_org(request)
    payroll = get_object_or_404(Payroll, pk=pk, organization=org)

    if payroll.status != 'draft':
        messages.error(request, 'Only draft payrolls can be processed.')
        return redirect('payroll_detail', pk=pk)

    payroll.status = 'processed'
    payroll.save(update_fields=['status', 'updated_at'])
    messages.success(request, f'Payroll {payroll.payroll_number} marked as processed.')
    return redirect('payroll_detail', pk=pk)


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def payroll_mark_paid(request, pk):
    """Mark a processed payroll as paid."""
    org = get_org(request)
    payroll = get_object_or_404(Payroll, pk=pk, organization=org)

    if payroll.status != 'processed':
        messages.error(request, 'Only processed payrolls can be marked as paid.')
        return redirect('payroll_detail', pk=pk)

    payroll.status = 'paid'
    payroll.payment_date = date.today()
    payroll.payment_method = request.POST.get('payment_method', 'Cash')
    payroll.save(update_fields=['status', 'payment_date', 'payment_method', 'updated_at'])
    messages.success(request, f'Payroll {payroll.payroll_number} marked as paid.')
    return redirect('payroll_detail', pk=pk)


@login_required(login_url='login')
@manager_or_admin_required
def payroll_payslip_print(request, pk):
    """Printable pay slip for admin."""
    org = get_org(request)
    payroll = get_object_or_404(Payroll, pk=pk, organization=org)

    components = payroll.components.all()
    earnings = components.filter(component_type='earning')
    deductions = components.filter(component_type='deduction')

    context = {
        'payroll': payroll,
        'earnings': earnings,
        'deductions': deductions,
        'staff': payroll.staff,
        'org': org,
    }
    return render(request, 'management/staff_my_payslip.html', context)


# ─── Accounts & Expenses ─────────────────────────────────────────────────────

@login_required(login_url='login')
@manager_or_admin_required
def accounts_overview(request):
    """Balance sheet / financial report with date filters."""
    org = get_org(request)
    today = date.today()

    # Date range filter
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
    except ValueError:
        start_date = None
    try:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except ValueError:
        end_date = None

    # --- INCOME: Fee Payments ---
    fee_qs = FeePayment.objects.filter(organization=org)
    if start_date:
        fee_qs = fee_qs.filter(payment_date__gte=start_date)
    if end_date:
        fee_qs = fee_qs.filter(payment_date__lte=end_date)

    fee_total = fee_qs.aggregate(total=Sum('amount'))['total'] or 0
    fee_by_method = fee_qs.values('payment_method').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('payment_method')

    # --- EXPENSES: General Expenses ---
    expense_qs = Expense.objects.filter(organization=org)
    if start_date:
        expense_qs = expense_qs.filter(expense_date__gte=start_date)
    if end_date:
        expense_qs = expense_qs.filter(expense_date__lte=end_date)

    expense_total = expense_qs.aggregate(total=Sum('amount'))['total'] or 0
    expense_by_method = expense_qs.values('payment_method').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('payment_method')
    expense_by_category = expense_qs.values('category').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')

    # Map category keys to display names
    category_display = dict(Expense.CATEGORY_CHOICES)
    for item in expense_by_category:
        item['category_display'] = category_display.get(item['category'], item['category'])

    # --- EXPENSES: Payroll (paid only) ---
    payroll_qs = Payroll.objects.filter(organization=org, status='paid')
    if start_date:
        payroll_qs = payroll_qs.filter(payment_date__gte=start_date)
    if end_date:
        payroll_qs = payroll_qs.filter(payment_date__lte=end_date)

    payroll_total = payroll_qs.aggregate(total=Sum('net_salary'))['total'] or 0
    payroll_by_method = payroll_qs.values('payment_method').annotate(
        total=Sum('net_salary'), count=Count('id')
    ).order_by('payment_method')

    # --- TOTALS ---
    total_expense = expense_total + payroll_total
    net_balance = fee_total - total_expense

    # --- Method-wise summary (all money in and out) ---
    methods = ['Cash', 'Bank Transfer', 'Online', 'UPI']
    method_summary = []
    fee_method_map = {item['payment_method']: item for item in fee_by_method}
    expense_method_map = {item['payment_method']: item for item in expense_by_method}
    payroll_method_map = {item['payment_method']: item for item in payroll_by_method}

    for method in methods:
        income = (fee_method_map.get(method, {}).get('total') or 0)
        exp = (expense_method_map.get(method, {}).get('total') or 0)
        pay = (payroll_method_map.get(method, {}).get('total') or 0)
        method_summary.append({
            'method': method,
            'income': income,
            'expense': exp + pay,
            'net': income - exp - pay,
        })

    # Recent transactions
    recent_fees = fee_qs.select_related('student', 'batch__course').order_by('-payment_date', '-created_at')[:5]
    recent_expenses = expense_qs.order_by('-expense_date', '-created_at')[:5]
    recent_payrolls = payroll_qs.select_related('staff').order_by('-payment_date', '-created_at')[:5]

    context = {
        'fee_total': fee_total,
        'expense_total': expense_total,
        'payroll_total': payroll_total,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'fee_by_method': fee_by_method,
        'expense_by_method': expense_by_method,
        'expense_by_category': expense_by_category,
        'payroll_by_method': payroll_by_method,
        'method_summary': method_summary,
        'recent_fees': recent_fees,
        'recent_expenses': recent_expenses,
        'recent_payrolls': recent_payrolls,
        'date_from': date_from,
        'date_to': date_to,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'management/accounts_overview.html', context)


@login_required(login_url='login')
@manager_or_admin_required
def expense_list(request):
    """List all expenses with search and filters."""
    org = get_org(request)
    expenses_qs = Expense.objects.filter(organization=org)

    search_query = request.GET.get('q', '').strip()
    if search_query:
        expenses_qs = expenses_qs.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(reference_number__icontains=search_query)
        )

    selected_category = request.GET.get('category', '').strip()
    if selected_category:
        expenses_qs = expenses_qs.filter(category=selected_category)

    selected_method = request.GET.get('method', '').strip()
    if selected_method:
        expenses_qs = expenses_qs.filter(payment_method=selected_method)

    paginator = Paginator(expenses_qs, 20)
    page_number = request.GET.get('page')
    expenses = paginator.get_page(page_number)
    return render(request, 'management/expense_list.html', {
        'expenses': expenses,
        'search_query': search_query,
        'selected_category': selected_category,
        'selected_method': selected_method,
        'category_choices': Expense.CATEGORY_CHOICES,
    })


@login_required(login_url='login')
@manager_or_admin_required
def expense_add(request):
    org = get_org(request)
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.organization = org
            expense.created_by = request.user
            expense.save()
            messages.success(request, f'Expense "{expense.title}" recorded!')
            return redirect('expense_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExpenseForm()
    return render(request, 'management/expense_form.html', {'form': form, 'action': 'Add'})


@login_required(login_url='login')
@manager_or_admin_required
def expense_edit(request, pk):
    org = get_org(request)
    expense = get_object_or_404(Expense, pk=pk, organization=org)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, f'Expense "{expense.title}" updated!')
            return redirect('expense_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'management/expense_form.html', {'form': form, 'action': 'Edit'})


@login_required(login_url='login')
@manager_or_admin_required
@require_POST
def expense_delete(request, pk):
    org = get_org(request)
    expense = get_object_or_404(Expense, pk=pk, organization=org)
    title = expense.title
    expense.delete()
    messages.success(request, f'Expense "{title}" deleted.')
    return redirect('expense_list')
